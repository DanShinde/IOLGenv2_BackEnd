"""
Builds the Summary Report worksheet matching the reference template.

Layout:
- "TEST SCORE" title with overall completion %
- "TEST CASE SUMMARY" heading with "Last Updated" date placed beside it on the right (Column H & I)
- "TEST CASE SUMMARY" table with columns:
  Feature/Module | Total Test Cases | Pass (green) | Failed (red) | Pending (orange) | Percentage (gray)
- TOTAL row at bottom
- Completion and Distribution charts side-by-side starting at Row 13:
  - Bar chart at A13: Sage green bars with blue borders
  - Pie chart at H13: Centered, non-overlapping labels

All values use COUNTIF formulas referencing section sheets for live updates.
"""
from datetime import date

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from .styles import (
    FONT_TITLE, FONT_SUBTITLE, FONT_SUMMARY_HEADER, FONT_TOTAL,
    FONT_LABEL, FONT_VALUE, FONT_PERCENTAGE,
    FILL_GREEN, FILL_RED, FILL_YELLOW_ORANGE, FILL_ORANGE, FILL_MEDIUM_GRAY,
    FILL_LIGHT_GRAY, BLACK,
    THIN_BORDER, ALIGN_CENTER, ALIGN_CENTER_NOWRAP, ALIGN_LEFT_NOWRAP,
    SUMMARY_COL_WIDTHS,
)


def build_summary_sheet(
    ws: Worksheet,
    section_names: list[str],
    section_data_ranges: dict[str, tuple[int, int]],
    protect_sheets: bool = True,
) -> None:
    """
    Build the Summary Report worksheet with live formulas and charts.

    Args:
        ws: The Summary worksheet
        section_names: List of section names (matching sheet names)
        section_data_ranges: Dict mapping section_name -> (first_data_row, last_data_row)
                           for building COUNTIF formulas. These are the row ranges
                           where Result values (column F) live in each section sheet.
    """
    # Enable grid lines visibility
    ws.views.sheetView[0].showGridLines = True

    # ─── Column Widths ───────────────────────────────────────────────────
    for col_letter, width in SUMMARY_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ─── Title Block ─────────────────────────────────────────────────────
    # Row 1: "TEST SCORE" title
    cell = ws.cell(row=1, column=2, value="TEST SCORE")
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER

    # Row 2: Overall percentage (formula)
    ws.cell(row=2, column=2).alignment = ALIGN_CENTER
    ws.cell(row=2, column=2).font = FONT_SUBTITLE

    # ─── "TEST CASE SUMMARY" heading + Last Updated ──────────────────────
    row_heading = 5
    
    # Merge A5:F5 for heading title
    cell = ws.cell(row=row_heading, column=1, value="TEST CASE SUMMARY")
    cell.font = FONT_SUBTITLE
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(
        start_row=row_heading, start_column=1,
        end_row=row_heading, end_column=6
    )

    # Last Updated Label in Column H (8) and Value in Column I (9)
    lbl_cell = ws.cell(row=row_heading, column=8, value="Last Updated")
    lbl_cell.font = Font(name="Calibri", size=11, bold=True, color=BLACK)
    lbl_cell.alignment = Alignment(horizontal="right", vertical="center")

    val_cell = ws.cell(row=row_heading, column=9, value=date.today().strftime("%d-%m-%Y"))
    val_cell.font = Font(name="Calibri", size=11, color=BLACK)
    val_cell.alignment = Alignment(horizontal="left", vertical="center")

    # ─── Column Headers (Row 6) ──────────────────────────────────────────
    header_row = 6
    headers = [
        ("Feature / Module", None),
        ("Total Test cases", None),
        ("Pass", FILL_GREEN),
        ("Failed", FILL_RED),
        ("Pending", FILL_YELLOW_ORANGE),
        ("Percentage", FILL_MEDIUM_GRAY),
    ]

    for col_idx, (header, fill) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = FONT_SUMMARY_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        else:
            cell.fill = FILL_MEDIUM_GRAY

    # ─── Section Rows (with formulas) ────────────────────────────────────
    data_start_row = header_row + 1

    for idx, section_name in enumerate(section_names):
        row = data_start_row + idx
        sheet_name = _safe_sheet_name(section_name)

        # Get the data range for this section
        first_row, last_row = section_data_ranges.get(
            section_name, (7, 100)
        )
        result_range = f"'{sheet_name}'!$F${first_row}:$F${last_row}"

        # Column A: Feature / Module
        cell = ws.cell(row=row, column=1, value=section_name)
        cell.font = Font(name="Calibri", size=11, color=BLACK)
        cell.alignment = ALIGN_LEFT_NOWRAP
        cell.border = THIN_BORDER

        # Column B: Total Test Cases
        cell = ws.cell(
            row=row, column=2,
            value=f'=COUNTA({result_range})'
        )
        cell.font = Font(name="Calibri", size=11, color=BLACK)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        # Column C: Pass
        cell = ws.cell(
            row=row, column=3,
            value=f'=COUNTIF({result_range},"Pass")'
        )
        cell.font = Font(name="Calibri", size=11, color=BLACK)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        # Column D: Failed
        cell = ws.cell(
            row=row, column=4,
            value=f'=COUNTIF({result_range},"Fail")'
        )
        cell.font = Font(name="Calibri", size=11, color=BLACK)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        # Column E: Pending
        cell = ws.cell(
            row=row, column=5,
            value=f'=COUNTIF({result_range},"Pending")'
        )
        cell.font = Font(name="Calibri", size=11, color=BLACK)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        # Column F: Percentage = (Pass + Failed) / Total * 100
        cell = ws.cell(
            row=row, column=6,
            value=f'=IF(B{row}>0, (C{row}+D{row})/B{row}*100, 0)'
        )
        cell.font = FONT_PERCENTAGE
        cell.number_format = '0.00'
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # ─── TOTAL Row ───────────────────────────────────────────────────────
    total_row = data_start_row + len(section_names) + 1  # +1 for blank row
    last_data_row = data_start_row + len(section_names) - 1

    cell = ws.cell(row=total_row, column=1, value="TOTAL")
    cell.font = FONT_TOTAL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    cell.fill = FILL_LIGHT_GRAY

    for col in range(2, 6):
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=total_row, column=col,
            value=f'=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})'
        )
        cell.font = FONT_TOTAL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        cell.fill = FILL_LIGHT_GRAY

    # Percentage total = Total (Pass + Failed) / Total Test Cases * 100
    cell = ws.cell(
        row=total_row, column=6,
        value=f'=IF(B{total_row}>0, (C{total_row}+D{total_row})/B{total_row}*100, 0)'
    )
    cell.font = FONT_TOTAL
    cell.number_format = '0.00'
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    cell.fill = FILL_LIGHT_GRAY

    # ─── Overall % in Title Block ────────────────────────────────────────
    ws.cell(
        row=2, column=2,
        value=f'=IF(B{total_row}>0, ROUND((C{total_row}+D{total_row})/B{total_row}*100, 0), 0) & "%"'
    )

    # ─── Charts ──────────────────────────────────────────────────────────
    _add_bar_chart(ws, data_start_row, last_data_row, section_names)
    _add_pie_chart(ws, total_row)

    # Enable sheet protection
    if protect_sheets:
        ws.protection.sheet = True


def _add_pie_chart(ws: Worksheet, total_row: int) -> None:
    """Add a centered pie chart showing Pass, Failed, and Pending distribution."""
    chart = PieChart()
    chart.title = "Test Result Distribution"
    chart.style = 10
    chart.width = 14
    chart.height = 9

    # Labels: Pass, Failed, Pending (Columns C to E, row 6)
    labels = Reference(ws, min_col=3, max_col=5, min_row=6)
    # Data: totals from the TOTAL row
    data = Reference(ws, min_col=3, max_col=5, min_row=total_row)

    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(labels)

    # Style the slices
    chart.series[0].data_points = []

    # Green for Pass (idx 0)
    pt0 = DataPoint(idx=0)
    pt0.graphicalProperties.solidFill = "28A745"
    chart.series[0].data_points.append(pt0)

    # Red for Failed (idx 1)
    pt1 = DataPoint(idx=1)
    pt1.graphicalProperties.solidFill = "C00000"
    chart.series[0].data_points.append(pt1)

    # Orange for Pending (idx 2)
    pt2 = DataPoint(idx=2)
    pt2.graphicalProperties.solidFill = "FFA500"
    chart.series[0].data_points.append(pt2)

    # Configure data labels cleanly to prevent overlapping
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False

    # Legend position
    chart.legend.position = "r"

    # Add at column H, row 13 (side-by-side with bar chart)
    ws.add_chart(chart, "H13")


def _add_bar_chart(
    ws: Worksheet,
    data_start_row: int,
    last_data_row: int,
    section_names: list[str],
) -> None:
    """Add a bar chart showing completion percentage by section."""
    if not section_names:
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = "Completion % by Zone"
    chart.y_axis.title = "Percentage (%)"
    chart.x_axis.title = "Zone"
    chart.style = 10
    chart.width = 16
    chart.height = 9

    # Data: Percentage column (F)
    data = Reference(ws, min_col=6, min_row=data_start_row, max_row=last_data_row)
    # Categories: Section names (column A)
    cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=last_data_row)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.shape = 4
    chart.legend = None # Only single series, legend not needed

    # Color the bars to match second reference screenshot (light sage fill with blue border)
    if chart.series:
        series = chart.series[0]
        series.graphicalProperties.solidFill = "E2EFDA"
        series.graphicalProperties.line.solidFill = "5B9BD5"
        series.graphicalProperties.line.width = 12700 # ~1pt in openpyxl units

    # Add at column A, row 13
    ws.add_chart(chart, "A13")


def _safe_sheet_name(name: str) -> str:
    """Ensure sheet name is valid for Excel (max 31 chars, no special chars)."""
    # Remove characters not allowed in sheet names
    invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
    safe = name
    for char in invalid_chars:
        safe = safe.replace(char, "_")
    return safe[:31]
