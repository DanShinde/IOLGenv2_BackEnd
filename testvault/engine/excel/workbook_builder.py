"""
Orchestrates the full Excel workbook build.

Updated to use Section objects directly with the engine.
"""
from pathlib import Path
from openpyxl import Workbook

from ..dataclasses import Project, Section, GeneratedTestCase
from ..test_case_engine import TestCaseEngine
from .section_sheet import build_section_sheet
from .summary_sheet import build_summary_sheet


class WorkbookBuilder:
    """Builds the complete Excel workbook from a project configuration."""

    def __init__(self, engine: TestCaseEngine):
        self.engine = engine

    def build(self, project: Project, output_path: str, protect_sheets: bool = True, view_only: bool = False) -> dict:
        """
        Build the complete workbook and save to disk.

        Returns dict with generation statistics.
        """
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        if not view_only:
            ws_summary = wb.create_sheet(title="Summary Report")

        section_data_ranges: dict[str, tuple[int, int]] = {}
        section_stats: dict[str, int] = {}
        all_section_names: list[str] = []

        custom_tcs_by_section = []

        for section in project.sections:
            if not section.selected_groups and not (section.section_type not in ("conveyor", "vrc") and section.project_only_test_cases):
                continue

            # Generate standard test cases only
            std_test_cases = self.engine.generate_test_cases(section, exclude_custom=True)

            if std_test_cases:
                sheet_name = self._safe_sheet_name(section.name)
                ws = wb.create_sheet(title=sheet_name)

                data_row_count = build_section_sheet(
                    ws, project.info, section.name, std_test_cases, protect_sheets=protect_sheets, view_only=view_only
                )

                first_data_row, last_data_row = self._calculate_data_range(std_test_cases, view_only=view_only)

                section_data_ranges[section.name] = (first_data_row, last_data_row)
                section_stats[section.name] = data_row_count
                all_section_names.append(section.name)

            # Collect custom test cases from this active section
            sec_custom = self.engine.generate_test_cases(section, only_custom=True)
            if sec_custom:
                custom_tcs_by_section.extend(sec_custom)

        # Build Custom Test Cases sheet at the end if custom test cases are present
        if custom_tcs_by_section:
            # Group by category to maintain ordering and clean numbering
            custom_grouped = {}
            for tc in custom_tcs_by_section:
                if tc.category not in custom_grouped:
                    custom_grouped[tc.category] = []
                # Deduplicate by test case name
                if not any(x.test_case_name == tc.test_case_name for x in custom_grouped[tc.category]):
                    custom_grouped[tc.category].append(tc)

            ordered_custom_tcs = []
            for cat_idx, cat_name in enumerate(custom_grouped.keys(), start=1):
                for tc_idx, tc in enumerate(custom_grouped[cat_name], start=1):
                    tc.sr_no = f"{cat_idx}.{tc_idx}"
                    ordered_custom_tcs.append(tc)

            if ordered_custom_tcs:
                ws_custom = wb.create_sheet(title="Custom Test Cases")
                data_row_count = build_section_sheet(
                    ws_custom, project.info, "Custom Test Cases", ordered_custom_tcs, protect_sheets=protect_sheets, view_only=view_only
                )
                first_data_row, last_data_row = self._calculate_data_range(ordered_custom_tcs, view_only=view_only)
                section_data_ranges["Custom Test Cases"] = (first_data_row, last_data_row)
                section_stats["Custom Test Cases"] = data_row_count
                all_section_names.append("Custom Test Cases")

        if all_section_names and not view_only:
            build_summary_sheet(ws_summary, all_section_names, section_data_ranges, protect_sheets=protect_sheets)

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))

        return {
            "total_test_cases": sum(section_stats.values()),
            "sections": section_stats,
            "output_path": str(path),
        }

    def _calculate_data_range(
        self, test_cases: list[GeneratedTestCase], view_only: bool = False
    ) -> tuple[int, int]:
        """Calculate first and last data row numbers in the section sheet."""
        row_offset = 4 if view_only else 0
        start_row = 7 + row_offset
        if not test_cases:
            return (start_row, start_row)

        current_row = start_row
        first_data_row = None
        last_data_row = start_row
        current_category = None

        for tc in test_cases:
            if tc.category != current_category:
                current_category = tc.category
                current_row += 1  # Category header row

            if first_data_row is None:
                first_data_row = current_row
            last_data_row = current_row
            current_row += 1

        return (first_data_row or start_row, last_data_row)

    def _safe_sheet_name(self, name: str) -> str:
        """Ensure sheet name is valid for Excel."""
        invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
        safe = name
        for char in invalid_chars:
            safe = safe.replace(char, "_")
        return safe[:31]
