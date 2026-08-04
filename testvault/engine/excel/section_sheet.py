"""
Builds individual section worksheets matching the reference template.

Layout (matching reference screenshots):
- Rows 1-4: Project metadata in paired columns
- Row 5-6: Column headers (light gray background, black text, merged cells)
- Row 7+: Test cases grouped under category headers (dark maroon background, white text)
- Results column: Data validation dropdown (Pass, Fail, Pending, In Progress)
- Conditional formatting: Red fill when Result = "Fail"
"""
from pathlib import Path
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, Protection

from ..dataclasses import ProjectInfo, GeneratedTestCase
from .styles import (
    FONT_HEADER, FONT_LABEL, FONT_VALUE, FONT_VALUE_ITALIC, FONT_CATEGORY, FONT_DATA,
    FILL_HEADER, FILL_MEDIUM_GRAY, FILL_LIGHT_RED, FILL_LIGHT_GREEN,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER_NOWRAP,
    SECTION_COL_WIDTHS,
)


def build_section_sheet(
    ws: Worksheet,
    project_info: ProjectInfo,
    section_name: str,
    test_cases: list[GeneratedTestCase],
    protect_sheets: bool = True,
    view_only: bool = False,
) -> int:
    """
    Build a section worksheet with project info, headers, and test cases.

    Args:
        ws: The worksheet to populate
        project_info: Project metadata
        section_name: Name of this section
        test_cases: List of generated test cases for this section

    Returns:
        The number of test case data rows written (excluding headers/category rows)
    """
    # Enable grid lines visibility
    ws.views.sheetView[0].showGridLines = True

    # ─── Set Column Widths ────────────────────────────────────────────────
    for col_letter, width in SECTION_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    row_offset = 4 if view_only else 0

    # ─── Write Logo Banner (Row 1-4) if view_only ────────────────────────
    if view_only:
        ws.merge_cells("A1:G4")
        for r in range(1, 5):
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.border = THIN_BORDER
        logo_path = Path(__file__).resolve().parent.parent.parent / "static" / "testvault" / "assets" / "company_logo.png"
        if logo_path.exists():
            img = Image(str(logo_path))
            img.width = 320
            img.height = 40
            ws.add_image(img, "A1")

    # ─── Project Info Rows (1-4 or 5-8) ───────────────────────────────────
    _write_project_info(ws, project_info, section_name, view_only=view_only)

    # ─── Column Headers (Row 5 & 6 Merged) ────────────────────────────────
    # Set row heights
    ws.row_dimensions[5 + row_offset].height = 20
    ws.row_dimensions[6 + row_offset].height = 20

    # Headers definition
    headers = [
        (f"A{5+row_offset}", f"A{6+row_offset}", "Sr.\nNo."),
        (f"B{5+row_offset}", f"B{6+row_offset}", "Test Case"),
        (f"C{5+row_offset}", f"C{6+row_offset}", "Pre Requ. - State"),
        (f"D{5+row_offset}", f"D{6+row_offset}", "Action"),
        (f"E{5+row_offset}", f"E{6+row_offset}", "Expected Result"),
        (f"F{5+row_offset}", f"F{6+row_offset}", "Results\n(Pass / Fail)"),
        (f"G{5+row_offset}", f"G{6+row_offset}", "Observation\n(Checked by Validator)"),
    ]

    # Pre-fill all cells in header rows with borders and background fills
    for row in (5 + row_offset, 6 + row_offset):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = FONT_HEADER
            cell.fill = FILL_MEDIUM_GRAY
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER

    # Perform merges and set values
    for start_ref, end_ref, text in headers:
        ws.merge_cells(f"{start_ref}:{end_ref}")
        ws[start_ref] = text

    # ─── Freeze Panes ────────────────────────────────────────────────────
    ws.freeze_panes = f"A{7+row_offset}"

    # ─── Test Case Data ──────────────────────────────────────────────────
    current_row = 7 + row_offset
    current_category = None
    data_row_count = 0

    for tc in test_cases:
        # Write category header row if new category
        if tc.category != current_category:
            current_category = tc.category
            # Find category index from sr_no (e.g., "1.1" -> "1")
            cat_index = tc.sr_no.split(".")[0] if "." in tc.sr_no else "1"
            cat_label = f"{cat_index}. {tc.category}"

            # Set height for category row
            ws.row_dimensions[current_row].height = 20

            # Set value and formatting for all cells in the category row
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = FILL_HEADER
                cell.border = THIN_BORDER
                if col == 1:
                    cell.value = cat_label
                    cell.font = FONT_CATEGORY
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Merge across all columns for category header
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=7
            )
            current_row += 1

        # Write test case data row
        row_data = [
            tc.sr_no,
            tc.test_case_name,
            tc.pre_required_state,
            tc.action,
            tc.expected_result,
            tc.result,
            tc.observation,
        ]
        # Let Excel auto-fit the height of the row by not setting height explicitly

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = FONT_DATA
            cell.border = THIN_BORDER
            if col_idx == 1:  # Sr. No. centered
                cell.alignment = ALIGN_CENTER
            elif col_idx == 6:  # Results centered
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT

            # Unlock Results (Column F/6) and Observation (Column G/7) cells only if not view_only
            if not view_only and col_idx in (6, 7):
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)

        current_row += 1
        data_row_count += 1

    # ─── Data Validation for Results Column ──────────────────────────────
    if data_row_count > 0:
        dv = DataValidation(
            type="list",
            formula1='"Pass,Fail,Pending"',
            allow_blank=True,
        )
        dv.error = "Please select Pass, Fail, or Pending"
        dv.errorTitle = "Invalid Result"
        dv.prompt = "Select test result"
        dv.promptTitle = "Result"
        # Apply to all potential result cells
        _apply_validation_to_data_rows(ws, dv, test_cases, view_only=view_only)
        ws.add_data_validation(dv)

    # ─── Conditional Formatting ──────────────────────────────────────────
    # Red fill for Fail, Green fill for Pass results
    last_row = current_row - 1
    if last_row >= 7 + row_offset:
        result_range = f"A{7+row_offset}:G{last_row}"
        ws.conditional_formatting.add(
            result_range,
            CellIsRule(
                operator="equal",
                formula=['"Fail"'],
                fill=FILL_LIGHT_RED,
            ),
        )
        ws.conditional_formatting.add(
            result_range,
            CellIsRule(
                operator="equal",
                formula=['"Pass"'],
                fill=FILL_LIGHT_GREEN,
            ),
        )

    # Enable sheet protection to enforce locked cells
    if protect_sheets:
        ws.protection.sheet = True

    return data_row_count


def _write_project_info(
    ws: Worksheet, info: ProjectInfo, section_name: str, view_only: bool = False
) -> None:
    """Write project metadata in rows 1-4 or 5-8 using merged paired column layout."""
    row_offset = 4 if view_only else 0
    # Set metadata row heights to 25 to fit size 16 fonts nicely
    for row in range(1 + row_offset, 5 + row_offset):
        ws.row_dimensions[row].height = 25

    # Ensure all cells in metadata block have thin border and standard alignment
    for row in range(1 + row_offset, 5 + row_offset):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Perform merges for labels and values
    # Left side merges
    for row in range(1 + row_offset, 5 + row_offset):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2) # A:B for labels
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4) # C:D for values

    # Right side merges for values (F:G)
    for row in range(1 + row_offset, 4 + row_offset):
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7) # F:G for values

    # Write values and apply fonts
    # Row 1
    ws.cell(row=1 + row_offset, column=1, value="Project Code / Title:").font = FONT_LABEL
    ws.cell(row=1 + row_offset, column=3, value=info.project_code).font = FONT_VALUE
    ws.cell(row=1 + row_offset, column=5, value="Date of Validation:").font = FONT_LABEL
    ws.cell(row=1 + row_offset, column=6, value=info.date_of_validation).font = FONT_VALUE

    # Row 2
    if info.validator_type == "Self":
        validator_val = f"Self- {info.done_by}"
    else:
        validator_val = f"{info.validator_type}- {info.validator_name}"

    ws.cell(row=2 + row_offset, column=1, value="Zone Name.:").font = FONT_LABEL
    ws.cell(row=2 + row_offset, column=3, value=section_name).font = FONT_VALUE
    ws.cell(row=2 + row_offset, column=5, value="Self/ Internal Validator/External Validator").font = FONT_LABEL
    ws.cell(row=2 + row_offset, column=6, value=validator_val).font = FONT_VALUE

    # Row 3
    ws.cell(row=3 + row_offset, column=1, value="Customer Name").font = FONT_LABEL
    ws.cell(row=3 + row_offset, column=3, value=info.customer_name).font = FONT_VALUE
    ws.cell(row=3 + row_offset, column=5, value="Testing Phase (Emulation/FAT/SAT):").font = FONT_LABEL
    ws.cell(row=3 + row_offset, column=6, value=info.testing_phase).font = FONT_VALUE

    # Row 4
    ws.cell(row=4 + row_offset, column=1, value="Test Cases Prepared By:").font = FONT_LABEL
    ws.cell(row=4 + row_offset, column=3, value=info.done_by).font = FONT_VALUE

    # Footer logic: NOT present if view_only is True
    if not view_only:
        ws.merge_cells(start_row=4, start_column=5, end_row=4, end_column=7)
        footer_cell = ws.cell(row=4, column=5, value="Made by EFS TestCaseGenerator, cannot be edited.")
        footer_cell.font = Font(name="Calibri", size=11, italic=True, color="FF0000")
        footer_cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_validation_to_data_rows(
    ws: Worksheet, dv: DataValidation, test_cases: list[GeneratedTestCase], view_only: bool = False
) -> None:
    """Apply data validation only to actual data rows (not category headers)."""
    row_offset = 4 if view_only else 0
    # Walk through the sheet to find data rows starting at row 7 + offset
    row = 7 + row_offset
    current_category = None
    for tc in test_cases:
        if tc.category != current_category:
            current_category = tc.category
            row += 1  # Skip category header row

        # Add this data row's Result cell to validation
        cell_ref = f"F{row}"
        dv.add(cell_ref)
        row += 1
