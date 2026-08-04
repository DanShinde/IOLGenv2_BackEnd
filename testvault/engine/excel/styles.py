"""
Excel cell styles matching the reference template screenshots.

Reference style:
- Section column headers: Dark maroon/red (#8B0000), white bold text
- Category group headers: Bold, dark red text, spanning full row
- Failed rows: Light red background
- Pass rows: Default white
- Project info: Bold labels, regular values
- Summary 'Pass' column: Green header
- Summary 'Pending' column: Orange/yellow header  
- Summary 'Percentage' column: Orange header
"""
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    NamedStyle,
)


# ─── Colors ───────────────────────────────────────────────────────────────────

# Header colors (matching reference screenshots)
MAROON = "8B0000"
DARK_RED = "800000"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "C00000"

# Summary column header colors
GREEN = "28A745"
ORANGE = "FF8C00"
YELLOW_ORANGE = "FFA500"
LIGHT_GRAY = "F2F2F2"
MEDIUM_GRAY = "D9D9D9"

# Row highlighting
LIGHT_RED = "FFCCCC"
LIGHT_GREEN = "CCFFCC"
LIGHT_YELLOW = "FFFFCC"

# ─── Fonts ────────────────────────────────────────────────────────────────────

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=BLACK)
FONT_LABEL = Font(name="Calibri", size=12, bold=True, color=BLACK)
FONT_VALUE = Font(name="Calibri", size=16, italic=True, color=BLACK)
FONT_VALUE_ITALIC = Font(name="Calibri", size=16, italic=True, color=BLACK)
FONT_CATEGORY = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_DATA = Font(name="Calibri", size=10, color=BLACK)
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=GREEN)
FONT_SUBTITLE = Font(name="Calibri", size=14, bold=True, color=BLACK)
FONT_SUMMARY_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_TOTAL = Font(name="Calibri", size=11, bold=True, color=BLACK)
FONT_PERCENTAGE = Font(name="Calibri", size=11, color=BLACK)

# ─── Fills ────────────────────────────────────────────────────────────────────

FILL_HEADER = PatternFill(start_color=MAROON, end_color=MAROON, fill_type="solid")
FILL_GREEN = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
FILL_RED = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
FILL_ORANGE = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
FILL_YELLOW_ORANGE = PatternFill(
    start_color=YELLOW_ORANGE, end_color=YELLOW_ORANGE, fill_type="solid"
)
FILL_LIGHT_GRAY = PatternFill(
    start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
)
FILL_MEDIUM_GRAY = PatternFill(
    start_color=MEDIUM_GRAY, end_color=MEDIUM_GRAY, fill_type="solid"
)
FILL_LIGHT_RED = PatternFill(
    start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid"
)
FILL_LIGHT_GREEN = PatternFill(
    start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"
)
FILL_LIGHT_YELLOW = PatternFill(
    start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid"
)
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# ─── Borders ──────────────────────────────────────────────────────────────────

THIN_SIDE = Side(style="thin", color=BLACK)
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)
NO_BORDER = Border()

# ─── Alignment ────────────────────────────────────────────────────────────────

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER_NOWRAP = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_NOWRAP = Alignment(horizontal="left", vertical="center")

# ─── Column Widths ────────────────────────────────────────────────────────────

# Section sheet column widths (matching reference)
SECTION_COL_WIDTHS = {
    "A": 8,    # Sr. No.
    "B": 30,   # Test Case
    "C": 35,   # Pre Requ. - State
    "D": 40,   # Action
    "E": 50,   # Expected Result
    "F": 12,   # Results
    "G": 35,   # Observation
}

# Summary sheet column widths
SUMMARY_COL_WIDTHS = {
    "A": 25,   # Feature / Module
    "B": 18,   # Total Test Cases
    "C": 12,   # Pass
    "D": 12,   # Failed
    "E": 12,   # Pending
    "F": 16,   # Percentage
    "G": 10,   # spacer
    "H": 15,   # Last Updated Label
    "I": 15,   # Last Updated Value
}
