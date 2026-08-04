import openpyxl
from pathlib import Path
from typing import Optional
from ..dataclasses import Project, Section

def import_template_to_project(
    project_data: dict, file_path_or_stream, target_selection: str, data_dir: Optional[Path] = None
) -> tuple[dict, str]:
    """
    Parse the uploaded Excel template and import test cases into the project.
    
    Returns:
        tuple: (updated_project_dict, success_message)
    """
    try:
        wb = openpyxl.load_workbook(file_path_or_stream, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to open Excel file: {str(e)}")
        
    if not wb.sheetnames:
        raise ValueError("Excel file contains no sheets.")
        
    ws = wb.active
    
    # 1. Validate template headers
    # Expected headers: Test Case Name, Prerequisites, Action / Procedure, Expected Result
    expected_headers = ["Test Case Name", "Prerequisites", "Action / Procedure", "Expected Result"]
    
    actual_headers = []
    for col_idx in range(1, 5):
        val = ws.cell(row=1, column=col_idx).value
        actual_headers.append(str(val).strip() if val is not None else "")
        
    # Match headers case-insensitively/stripped
    for idx, expected in enumerate(expected_headers):
        if idx >= len(actual_headers) or actual_headers[idx].lower().replace(" ", "") != expected.lower().replace(" ", ""):
            raise ValueError(
                f"Invalid template headers. Expected columns A-D: {', '.join(expected_headers)}. "
                f"Found: {', '.join(actual_headers[:4])}"
            )
            
    # 2. Parse sections and test cases
    parsed_sections = {} # name -> list of test case dicts
    active_section = None
    
    for row_idx in range(2, ws.max_row + 1):
        val_a = ws.cell(row=row_idx, column=1).value
        val_b = ws.cell(row=row_idx, column=2).value
        val_c = ws.cell(row=row_idx, column=3).value
        val_d = ws.cell(row=row_idx, column=4).value
        
        str_a = str(val_a).strip() if val_a is not None else ""
        str_b = str(val_b).strip() if val_b is not None else ""
        str_c = str(val_c).strip() if val_c is not None else ""
        str_d = str(val_d).strip() if val_d is not None else ""
        
        if not str_a and not str_b and not str_c and not str_d:
            # Fully empty row, skip
            continue
            
        # Check if this row is a merged cell representing a Section Name (Columns A:D merged)
        is_section_row = False
        for merged_range in ws.merged_cells.ranges:
            if row_idx >= merged_range.min_row and row_idx <= merged_range.max_row:
                if 1 >= merged_range.min_col and 1 <= merged_range.max_col:
                    is_section_row = True
                    break
        
        if is_section_row:
            if not str_a:
                continue
            active_section = str_a
            if active_section not in parsed_sections:
                parsed_sections[active_section] = []
            continue
            
        # Test Case Name row
        if not str_a:
            raise ValueError(f"Row {row_idx}: Test Case Name in column A cannot be empty.")
            
        if active_section is None:
            raise ValueError(f"Row {row_idx}: Test case '{str_a}' found before any Section Name row.")
            
        parsed_sections[active_section].append({
            "name": str_a,
            "pre_required_state": str_b,
            "action": str_c,
            "expected_result": str_d
        })
        
    if not parsed_sections:
        raise ValueError("No sections or test cases found in the uploaded template.")
        
    import json
    import time
    import random

    # 3. Apply import rules to test_groups.json
    if data_dir is None:
        data_dir = Path(__file__).parent.parent / "data"
    tg_file = data_dir / "test_groups.json"
    if not tg_file.exists():
        raise ValueError("test_groups.json not found in Data directory.")
        
    try:
        with open(tg_file, "r", encoding="utf-8") as f:
            tg_data = json.load(f)
    except Exception as e:
        raise ValueError(f"Failed to read test_groups.json: {str(e)}")
        
    categories = tg_data.setdefault("categories", [])
    
    total_imported = 0
    categories_created = 0
    clusters_created = 0
    clusters_appended = 0
    
    for sec_name, tcs in parsed_sections.items():
        if not tcs:
            continue
            
        # Find if a group matching the section name and target selection already exists
        existing_group = None
        existing_cat = None
        for cat in categories:
            for grp in cat.setdefault("groups", []):
                if grp.get("label", "").lower() == sec_name.lower() and grp.get("section_type", "").lower() == target_selection.lower():
                    existing_group = grp
                    existing_cat = cat
                    break
            if existing_group:
                break
                
        if existing_group:
            # Append test cases to existing group
            existing_names = {tc.get("name").lower() for tc in existing_group.setdefault("test_cases", [])}
            for tc in tcs:
                if tc["name"].lower() not in existing_names:
                    tc_obj = {
                        "name": tc["name"],
                        "pre_required_state": tc.get("pre_required_state", ""),
                        "action": tc.get("action", ""),
                        "expected_result": tc.get("expected_result", "")
                    }
                    existing_group["test_cases"].append(tc_obj)
                    existing_names.add(tc["name"].lower())
                    total_imported += 1
            clusters_appended += 1
        else:
            # Create a new group/cluster
            target_cat = None
            for cat in categories:
                if cat.get("name", "").lower() == sec_name.lower():
                    target_cat = cat
                    break
                    
            if not target_cat:
                cat_id = "cat_" + str(int(time.time())) + "_" + str(random.randint(10, 99))
                target_cat = {
                    "name": sec_name,
                    "id": cat_id,
                    "category_number": len(categories) + 1,
                    "groups": []
                }
                categories.append(target_cat)
                categories_created += 1
                
            group_id = "group_" + str(int(time.time() * 1000)) + "_" + str(random.randint(100, 999))
            tcs_to_add = []
            for tc in tcs:
                tc_obj = {
                    "name": tc["name"],
                    "pre_required_state": tc.get("pre_required_state", ""),
                    "action": tc.get("action", ""),
                    "expected_result": tc.get("expected_result", "")
                }
                tcs_to_add.append(tc_obj)
                total_imported += 1
                
            new_grp = {
                "id": group_id,
                "label": sec_name,
                "section_type": target_selection.lower(),
                "ui_type": "checkbox",
                "test_cases": tcs_to_add
            }
            target_cat.setdefault("groups", []).append(new_grp)
            clusters_created += 1
            
    # Save the updated test_groups.json
    try:
        with open(tg_file, "w", encoding="utf-8") as f:
            json.dump(tg_data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        raise ValueError(f"Failed to write test_groups.json: {str(e)}")
        
    msg = f"Import successful: Appended to {clusters_appended} clusters, created {clusters_created} clusters (across {categories_created} new categories), and imported {total_imported} test cases."
    return project_data, msg
