import json
import os
import tempfile
import uuid
from pathlib import Path

from django.contrib.auth import authenticate
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_GET, require_http_methods

from employees.models import Employee
from planner.models import Project as PlannerProject
from tracker.models import Project as TrackerProject

from .engine.changelog import APP_VERSION, CHANGELOG
from .engine.dataclasses import Project as EngineProject, Section
from .engine.excel.template_importer import import_template_to_project
from .engine.excel.workbook_builder import WorkbookBuilder
from .models import ReportSession, TestVaultProject
from .services import compute_project_stats, get_data_dir, get_engine

APP_DIR = Path(__file__).resolve().parent


# ─── SPA shell ──────────────────────────────────────────────────────────────

def index(request):
    return render(request, "testvault/index.html")


def report(request):
    return render(request, "testvault/report.html")


# ─── Reference data / test case generation (ported ~1:1 from Web/server.py) ─

@require_GET
def api_changelog(request):
    return JsonResponse({"version": APP_VERSION, "changelog": CHANGELOG})


@require_GET
def api_manual(request):
    manual_path = APP_DIR / "static" / "testvault" / "docs" / "EFS_TestCaseGenerator_User_Manual.pdf"
    if manual_path.exists():
        return FileResponse(open(manual_path, "rb"), content_type="application/pdf")
    return JsonResponse({"error": "User Manual not found"}, status=404)


@require_GET
def api_test_groups(request):
    return JsonResponse(get_engine()._test_groups_data)


@require_http_methods(["POST"])
def api_generate_test_cases(request):
    try:
        data = json.loads(request.body)
        section = Section.from_dict(data)
        tcs = get_engine().generate_test_cases(section)
        return JsonResponse({"test_cases": [tc.__dict__ for tc in tcs]})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_http_methods(["POST"])
def api_generate_workbook(request):
    try:
        data = json.loads(request.body)
        project_data = data.get("project")
        protect_sheets = data.get("protect_sheets", True)
        view_only = data.get("view_only", False)

        project = EngineProject.from_dict(project_data)
        builder = WorkbookBuilder(get_engine())

        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            builder.build(project, temp_path, protect_sheets=protect_sheets, view_only=view_only)

            code = project.info.project_code.strip() or "UNKNOWN_PROJECT"
            customer = project.info.customer_name.strip() or "UNKNOWN_CUSTOMER"

            active_types = []
            for s in project.sections:
                is_active = bool(s.selected_groups) or bool(s.project_only_test_cases)
                if is_active and s.section_type:
                    stype = s.section_type.strip()
                    if stype:
                        suffix_part = stype[:4].upper()
                        if suffix_part not in active_types:
                            active_types.append(suffix_part)
            suffix = "_".join(active_types) if active_types else "CONV"

            filename = f"{code}-{customer}-{suffix}-AUTO-TEST-CASES.xlsx"
            filename = "".join(c for c in filename if c not in '\\/:*?"<>|')

            with open(temp_path, "rb") as f:
                content = f.read()
            response = HttpResponse(
                content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_GET
def api_custom_test_cases(request):
    return JsonResponse({"custom_test_cases": get_engine()._load_permanent_custom_test_cases()})


@require_http_methods(["POST"])
def api_custom_test_cases_add(request):
    try:
        tc = json.loads(request.body)
        get_engine().add_permanent_custom_test_case(tc)
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_http_methods(["POST"])
def api_custom_test_cases_delete(request):
    try:
        data = json.loads(request.body)
        name = data.get("name")
        if name:
            get_engine().delete_permanent_custom_test_case(name)
            return JsonResponse({"success": True})
        return JsonResponse({"error": "Name field missing"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_http_methods(["POST"])
def api_edit_test_groups(request):
    try:
        data = json.loads(request.body)
        engine = get_engine()
        categories = data.get("categories")
        selections = data.get("selections")

        if categories is not None:
            engine._test_groups_data["categories"] = categories
        if selections is not None:
            engine._test_groups_data["selections"] = selections

        tg_file = engine.data_dir / "test_groups.json"
        with open(tg_file, "w", encoding="utf-8") as f:
            json.dump(engine._test_groups_data, f, indent=2, ensure_ascii=False)

        engine.reload_data()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# ─── Shareable report sessions (DB-backed, see models.ReportSession) ───────

@require_http_methods(["POST"])
def api_report_create(request):
    try:
        data = json.loads(request.body)
        if not data or "project" not in data:
            return JsonResponse({"error": "Missing project data"}, status=400)
        report_id = str(uuid.uuid4())[:8]
        ReportSession.objects.create(id=report_id, project_data=data["project"])
        return JsonResponse({"success": True, "report_id": report_id})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_http_methods(["POST"])
def api_report_update(request, report_id):
    session = ReportSession.objects.filter(id=report_id).first()
    if not session:
        return JsonResponse({"error": "Report session not found"}, status=404)
    try:
        data = json.loads(request.body)
        if not data or "project" not in data:
            return JsonResponse({"error": "Missing project data"}, status=400)
        session.project_data = data["project"]
        session.save(update_fields=["project_data", "updated_at"])
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_GET
def api_report_data(request, report_id):
    session = ReportSession.objects.filter(id=report_id).first()
    if not session:
        return JsonResponse({"error": "Report session not found"}, status=404)
    return JsonResponse({"project": session.project_data})


# ─── Template import/export ─────────────────────────────────────────────────

@require_GET
def api_template_download(request):
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    header_fill = PatternFill(start_color="6B3636", end_color="6B3636", fill_type="solid")
    zone_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", indent=1)

    thin_side = Side(style='thin', color='4A4A4A')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers = ["Test Case Name", "Prerequisites", "Action / Procedure", "Expected Result"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    zone_cell = ws.cell(row=2, column=1, value="Zone Name")
    zone_cell.font = white_bold
    zone_cell.fill = zone_fill
    zone_cell.alignment = left_align
    zone_cell.border = cell_border
    for col_idx in range(2, 5):
        c = ws.cell(row=2, column=col_idx)
        c.fill = zone_fill
        c.border = cell_border
    ws.row_dimensions[2].height = 24

    light_side = Side(style='thin', color='BFBFBF')
    light_border = Border(left=light_side, right=light_side, top=light_side, bottom=light_side)
    data_align = Alignment(vertical="top", wrap_text=True)
    for row_idx in range(3, 22):
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = light_border
            cell.alignment = data_align

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    response = HttpResponse(
        out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="EFS_Import_Template.xlsx"'
    return response


@require_http_methods(["POST"])
def api_template_upload(request):
    try:
        file = request.FILES.get("file")
        project_str = request.POST.get("project")
        target_selection = request.POST.get("target_selection", "conveyor")

        if not file:
            return JsonResponse({"error": "No file uploaded"}, status=400)
        if not project_str:
            return JsonResponse({"error": "No project state provided"}, status=400)

        project_data = json.loads(project_str)

        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        with os.fdopen(fd, "wb") as tmp:
            for chunk in file.chunks():
                tmp.write(chunk)

        try:
            engine = get_engine()
            updated_project, msg = import_template_to_project(
                project_data, tmp_path, target_selection, data_dir=get_data_dir()
            )
            engine.reload_data()
            return JsonResponse({
                "success": True,
                "project": updated_project,
                "message": msg,
                "test_groups": {"categories": engine.get_categories()},
            })
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# ─── Project CRUD (DB-backed, tied to tracker.Project + employees.Employee) ─

def _meta_from_request(data: dict, tv: TestVaultProject) -> None:
    """Apply the editable info-card fields onto a TestVaultProject instance in place."""
    if "zone_name" in data:
        tv.zone_name = data.get("zone_name") or ""
    if "prepared_by_id" in data:
        tv.prepared_by_id = data.get("prepared_by_id") or None
    if "date_of_validation" in data:
        tv.date_of_validation = data.get("date_of_validation") or ""
    if "validator_type" in data:
        tv.validator_type = data.get("validator_type") or "Self"
    if "validator_employee_id" in data:
        tv.validator_employee_id = data.get("validator_employee_id") or None
    if "validator_name" in data:
        tv.validator_name = data.get("validator_name") or ""
    if "testing_phase" in data:
        tv.testing_phase = data.get("testing_phase") or "Emulation"


@require_GET
def api_project_list(request):
    projects = TestVaultProject.objects.select_related("tracker_project").all()
    result = []
    for p in projects:
        result.append({
            "id": p.id,
            "project_code": p.project_code,
            "customer_name": p.customer_name,
            "zone_name": p.zone_name,
            "testing_phase": p.testing_phase,
            "updated_at": p.updated_at.isoformat(),
            "stats": compute_project_stats(p.sections),
        })
    return JsonResponse({"projects": result})


@require_http_methods(["POST"])
def api_project_create(request):
    try:
        data = json.loads(request.body)
        source = data.get("source")

        if source == "tracker":
            tracker_project = get_object_or_404(TrackerProject, pk=data.get("tracker_project_id"))
        elif source == "planner":
            planner_project = get_object_or_404(PlannerProject, pk=data.get("planner_project_id"))
            if not planner_project.tracker_project_id:
                return JsonResponse({
                    "error": "This Planner project isn't linked to a Tracker project yet. "
                             "Link it in Planner first, or pick a Tracker project directly."
                }, status=400)
            tracker_project = planner_project.tracker_project
        else:
            return JsonResponse({"error": "source must be 'tracker' or 'planner'"}, status=400)

        tv = TestVaultProject(tracker_project=tracker_project)
        _meta_from_request(data, tv)
        tv.created_by = request.user if request.user.is_authenticated else None
        tv.save()
        return JsonResponse({"success": True, "id": tv.id})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_GET
def api_project_detail(request, pk):
    tv = get_object_or_404(TestVaultProject, pk=pk)
    return JsonResponse({
        "id": tv.id,
        "project": tv.to_engine_dict(),
        "meta": {
            "tracker_project_id": tv.tracker_project_id,
            "zone_name": tv.zone_name,
            "prepared_by_id": tv.prepared_by_id,
            "date_of_validation": tv.date_of_validation,
            "validator_type": tv.validator_type,
            "validator_employee_id": tv.validator_employee_id,
            "validator_name": tv.validator_name,
            "testing_phase": tv.testing_phase,
        },
    })


@require_http_methods(["POST"])
def api_project_save(request, pk):
    tv = get_object_or_404(TestVaultProject, pk=pk)
    try:
        data = json.loads(request.body)
        project = data.get("project")
        if project is not None:
            tv.sections = project.get("sections", [])
            tv.custom_selection_types = project.get("custom_selection_types", [])
            tv.deleted_selection_types = project.get("deleted_selection_types", [])
        _meta_from_request(data.get("meta", {}), tv)
        tv.save()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_http_methods(["POST"])
def api_project_duplicate(request, pk):
    tv = get_object_or_404(TestVaultProject, pk=pk)
    clone = TestVaultProject.objects.create(
        tracker_project=tv.tracker_project,
        zone_name=f"{tv.zone_name} (Copy)" if tv.zone_name else "Copy",
        prepared_by=tv.prepared_by,
        date_of_validation=tv.date_of_validation,
        validator_type=tv.validator_type,
        validator_employee=tv.validator_employee,
        validator_name=tv.validator_name,
        testing_phase=tv.testing_phase,
        sections=tv.sections,
        custom_selection_types=tv.custom_selection_types,
        deleted_selection_types=tv.deleted_selection_types,
        created_by=request.user if request.user.is_authenticated else None,
    )
    return JsonResponse({"success": True, "id": clone.id})


@require_http_methods(["POST"])
def api_project_delete(request, pk):
    tv = get_object_or_404(TestVaultProject, pk=pk)
    try:
        data = json.loads(request.body)
    except Exception:
        data = {}
    password = data.get("password", "")
    if not request.user.is_authenticated or authenticate(username=request.user.username, password=password) is None:
        return JsonResponse({"error": "Incorrect password."}, status=403)
    tv.delete()
    return JsonResponse({"success": True})


# ─── Lookups (typeahead for project-create / prepared-by / validator pickers) ─

@require_GET
def api_lookup_tracker_projects(request):
    q = request.GET.get("q", "").strip()
    qs = TrackerProject.objects.all()
    if q:
        qs = qs.filter(code__icontains=q) | qs.filter(customer_name__icontains=q)
    qs = qs.order_by("code")[:25]
    return JsonResponse({"results": [{"id": p.id, "code": p.code, "customer_name": p.customer_name} for p in qs]})


@require_GET
def api_lookup_planner_projects(request):
    q = request.GET.get("q", "").strip()
    qs = PlannerProject.objects.all()
    if q:
        qs = qs.filter(project_id__icontains=q) | qs.filter(customer_name__icontains=q)
    qs = qs.order_by("project_id")[:25]
    return JsonResponse({"results": [
        {"id": p.id, "project_id": p.project_id, "customer_name": p.customer_name,
         "tracker_project_id": p.tracker_project_id}
        for p in qs
    ]})


@require_GET
def api_lookup_employees(request):
    q = request.GET.get("q", "").strip()
    qs = Employee.objects.filter(is_active=True)
    if q:
        qs = qs.filter(name__icontains=q)
    qs = qs.order_by("name")[:25]
    return JsonResponse({"results": [{"id": e.id, "name": e.name, "designation": e.designation} for e in qs]})
