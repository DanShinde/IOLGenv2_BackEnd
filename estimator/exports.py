"""Renders the estimate built by calculations.build_project_estimate() into Excel
(openpyxl) and PDF (reportlab). Both consume the exact same data structure so the two
formats -- and the on-screen report -- can never disagree with each other.
"""
from io import BytesIO

from django.http import HttpResponse
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

HEADER_FILL = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TOTAL_FILL = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
THIN_BORDER = Border(*(Side(style='thin', color='E2E8F0'),) * 4)


def _filename(project, ext):
    return f"{slugify(project.name) or 'project'}-estimate.{ext}"


# --------------------------------------------------------------------------- Excel

def render_project_report_excel(estimate):
    wb = Workbook()
    _write_summary_sheet(wb.active, estimate)
    for group in estimate['category_groups']:
        _write_breakdown_sheet(wb.create_sheet(group['label'][:31]), group)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_filename(estimate["project"], "xlsx")}"'
    return response


def _write_summary_sheet(ws, estimate):
    project = estimate['project']
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    ws['A1'] = f"Estimate: {project.name}"
    ws['A1'].font = Font(size=16, bold=True, color='1E293B')
    ws.merge_cells('A1:E1')

    # Headline figure: the final report is about man-days above all else.
    ws['A2'] = f"{float(estimate['grand_total_days']):.2f} man-days"
    ws['A2'].font = Font(size=22, bold=True, color='4F46E5')
    ws.merge_cells('A2:E2')
    ws['A3'] = f"@ {project.minutes_per_working_day} min/working day"
    ws['A3'].font = Font(italic=True, color='64748B')
    ws.merge_cells('A3:E3')

    ws['A4'] = f"Customer: {project.customer or '-'}   |   Complexity: {project.complexity.name} (x{project.complexity.multiplier})"
    ws['A4'].font = Font(italic=True, color='64748B')
    ws.merge_cells('A4:E4')

    row = 5
    for group in estimate['category_groups']:
        row += 1
        ws.cell(row=row, column=1, value=f"{group['label']}:").font = Font(bold=True, color='64748B')
        ws.cell(row=row, column=2, value=f"{float(group['total_days']):.2f} man-days").font = Font(bold=True, color='4F46E5')

    row += 2
    headers = ['Segment', 'Module Type', 'Count', 'Complexity', 'Row Total (min)', 'Row Total (man-days)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for r in estimate['rows']:
        ws.cell(row=row, column=1, value=r['segment'].name)
        ws.cell(row=row, column=2, value=r['module_type'].name)
        ws.cell(row=row, column=3, value=r['count'])
        ws.cell(row=row, column=4, value=f"{r['complexity'].name} (x{r['complexity'].multiplier})" if r['complexity'] else '-')
        ws.cell(row=row, column=5, value=float(r['row_total_minutes']))
        days_cell = ws.cell(row=row, column=6, value=round(float(r['row_total_days']), 2))
        days_cell.font = Font(bold=True, color='4F46E5')
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Grand Total').font = Font(bold=True)
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    ws.cell(row=row, column=2, value=f"{float(estimate['grand_total_minutes'])} min").fill = TOTAL_FILL
    ws.cell(row=row, column=3, value=f"{float(estimate['grand_total_hours']):.2f} hrs").fill = TOTAL_FILL
    ws.cell(row=row, column=4, value=f"{float(estimate['grand_total_days']):.2f} man-days").fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18


def _write_breakdown_sheet(ws, group):
    ws.sheet_view.showGridLines = False
    activities = group['activities']
    rows = group['rows']

    ws.cell(row=1, column=1, value='Segment').font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=2, value='Module Type').font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    ws.cell(row=1, column=3, value='Count').font = HEADER_FONT
    ws.cell(row=1, column=3).fill = HEADER_FILL

    for col, activity in enumerate(activities, start=4):
        cell = ws.cell(row=1, column=col, value=f"{activity.name} (man-days)")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 16

    days_col = len(activities) + 4
    row_total_col = days_col - 1
    ws.cell(row=1, column=row_total_col, value='Row Total (min)').font = HEADER_FONT
    ws.cell(row=1, column=row_total_col).fill = HEADER_FILL
    ws.cell(row=1, column=row_total_col).alignment = Alignment(horizontal='center', wrap_text=True)
    ws.cell(row=1, column=days_col, value='Row Total (man-days)').font = HEADER_FONT
    ws.cell(row=1, column=days_col).fill = HEADER_FILL
    ws.cell(row=1, column=days_col).alignment = Alignment(horizontal='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(row_total_col)].width = 16
    ws.column_dimensions[get_column_letter(days_col)].width = 18

    for row_offset, r in enumerate(rows):
        row = row_offset + 2
        ws.cell(row=row, column=1, value=r['segment'].name)
        ws.cell(row=row, column=2, value=r['module_type'].name)
        ws.cell(row=row, column=3, value=r['count'])
        for col, act_data in enumerate(r['activities'], start=4):
            cell = ws.cell(row=row, column=col, value=round(float(act_data['days']), 3))
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
        row_total_cell = ws.cell(row=row, column=row_total_col, value=round(float(r['row_total_minutes']), 1))
        row_total_cell.alignment = Alignment(horizontal='center')
        row_total_cell.font = Font(bold=True)
        days_cell = ws.cell(row=row, column=days_col, value=round(float(r['row_total_days']), 2))
        days_cell.alignment = Alignment(horizontal='center')
        days_cell.font = Font(bold=True, color='4F46E5')

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value='Activity Total').font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=2).fill = TOTAL_FILL
    ws.cell(row=total_row, column=3).fill = TOTAL_FILL
    for col, at in enumerate(group['activity_totals'], start=4):
        cell = ws.cell(row=total_row, column=col, value=round(float(at['total_days']), 3))
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=row_total_col, value=round(float(group['total_minutes']), 1)).font = Font(bold=True)
    ws.cell(row=total_row, column=row_total_col).fill = TOTAL_FILL
    ws.cell(row=total_row, column=days_col, value=round(float(group['total_days']), 2)).font = Font(bold=True, color='4F46E5')
    ws.cell(row=total_row, column=days_col).fill = TOTAL_FILL

    days_row = total_row + 2
    ws.cell(row=days_row, column=1, value=f"{group['label']} total").font = Font(bold=True, color='4F46E5')
    ws.cell(row=days_row, column=2, value=f"{float(group['total_days']):.2f} man-days").font = Font(bold=True, size=13, color='4F46E5')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.freeze_panes = 'D2'


# --------------------------------------------------------------------------- PDF

def render_project_report_pdf(estimate):
    project = estimate['project']
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Title'], textColor=colors.HexColor('#1E293B'))
    headline_style = ParagraphStyle('Headline', parent=styles['Title'], fontSize=32, leading=36,
                                     textColor=colors.HexColor('#4F46E5'), spaceBefore=4, spaceAfter=0)
    heading_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], textColor=colors.HexColor('#1E293B'),
                                    spaceBefore=14, spaceAfter=6)
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], textColor=colors.HexColor('#64748B'))

    story = [
        Paragraph(f"Estimate: {project.name}", title_style),
        # The final report is about man-days above all else -- lead with it.
        Paragraph(f"{estimate['grand_total_days']:.2f} man-days", headline_style),
        Paragraph(f"@ {estimate['minutes_per_day']} min/working day", meta_style),
        Spacer(1, 0.3 * cm),
        Paragraph(
            f"Customer: {project.customer or '-'} &nbsp;|&nbsp; "
            f"Complexity: {project.complexity.name} (x{project.complexity.multiplier})",
            meta_style,
        ),
        Spacer(1, 0.4 * cm),
    ]

    if estimate['warnings']:
        for w in estimate['warnings']:
            story.append(Paragraph(f"&#9888; {w}", ParagraphStyle('Warn', parent=meta_style, textColor=colors.HexColor('#D97706'))))
        story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph('Module Summary', heading_style))
    story.append(_module_summary_table(estimate))

    for group in estimate['category_groups']:
        story.append(Paragraph(
            f"{group['label']} &mdash; Activity-wise Breakdown (man-days) "
            f"<font color='#4F46E5'>({group['total_days']:.2f} man-days total)</font>",
            heading_style,
        ))
        story.extend(_activity_breakdown_tables(group))

    story.append(Spacer(1, 0.3 * cm))
    story.append(_grand_total_table(estimate))

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_filename(project, "pdf")}"'
    return response


def _module_summary_table(estimate):
    data = [['Segment', 'Module Type', 'Count', 'Complexity', 'Row Total (min)', 'Row Total (man-days)']]
    for r in estimate['rows']:
        data.append([
            r['segment'].name,
            r['module_type'].name,
            str(r['count']),
            f"{r['complexity'].name} (x{r['complexity'].multiplier})" if r['complexity'] else '-',
            f"{r['row_total_minutes']:.1f}",
            f"{r['row_total_days']:.2f}",
        ])
    table = Table(data, colWidths=[4 * cm, 5 * cm, 2 * cm, 4.5 * cm, 4 * cm, 4.5 * cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('FONTNAME', (5, 1), (5, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (5, 1), (5, -1), colors.HexColor('#4F46E5')),
    ]))
    return table


EMPLOYEES_PER_CHUNK = 6  # activity columns per PDF table chunk, so wide matrices still fit the page


def _activity_breakdown_tables(group):
    activities = group['activities']
    rows = group['rows']
    flowables = []

    last_chunk_start = max(0, ((len(activities) - 1) // EMPLOYEES_PER_CHUNK) * EMPLOYEES_PER_CHUNK) if activities else 0

    for chunk_start in range(0, len(activities) or 1, EMPLOYEES_PER_CHUNK):
        chunk = activities[chunk_start:chunk_start + EMPLOYEES_PER_CHUNK]
        is_last_chunk = chunk_start == last_chunk_start
        header = ['Segment', 'Module Type', 'Count'] + [f'{a.name} (days)' for a in chunk]
        if is_last_chunk:
            header += ['Row Total (min)', 'Row Total (man-days)']
        data = [header]
        for r in rows:
            act_by_id = {ad['activity'].id: ad for ad in r['activities']}
            row_vals = [r['segment'].name, r['module_type'].name, str(r['count'])]
            for a in chunk:
                row_vals.append(f"{act_by_id[a.id]['days']:.3f}")
            if is_last_chunk:
                row_vals += [f"{r['row_total_minutes']:.1f}", f"{r['row_total_days']:.2f}"]
            data.append(row_vals)

        totals_by_id = {at['activity'].id: at for at in group['activity_totals']}
        total_row = ['Activity Total', '', '']
        for a in chunk:
            total_row.append(f"{totals_by_id[a.id]['total_days']:.3f}")
        if is_last_chunk:
            total_row += [f"{group['total_minutes']:.1f}", f"{group['total_days']:.2f}"]
        data.append(total_row)

        extra_cols = 2 if is_last_chunk else 0
        col_widths = [3.5 * cm, 4.5 * cm, 2 * cm] + [((24 * cm) - 10 * cm - extra_cols * 3 * cm) / max(len(chunk), 1)] * len(chunk)
        if is_last_chunk:
            col_widths += [3 * cm, 3 * cm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EEF2FF')),
        ]
        if is_last_chunk:
            style_commands += [
                ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),
                ('TEXTCOLOR', (-1, 1), (-1, -1), colors.HexColor('#4F46E5')),
            ]
        table.setStyle(TableStyle(style_commands))
        flowables.append(table)
        flowables.append(Spacer(1, 0.4 * cm))

    return flowables


def _grand_total_table(estimate):
    data = [
        ['Man-days', 'Hours', 'Minutes'],
        [
            f"{estimate['grand_total_days']:.2f}",
            f"{estimate['grand_total_hours']:.2f}",
            f"{estimate['grand_total_minutes']:.1f}",
        ],
    ]
    table = Table(data, colWidths=[7 * cm] * 3)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#4F46E5')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    return table
