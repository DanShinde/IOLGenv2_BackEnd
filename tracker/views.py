from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.utils.dateparse import parse_date
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm

from employees.models import Employee
from .models import Stage, StageHistory, trackerSegment, StageRemark, ProjectUpdate, UpdateRemark, Project, ContactPerson, ProjectComment, SavedReportFilter

from django.db.models import Q, F, Sum, Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import date, timedelta, datetime

# Guards the Avg Delay / Avg Cycle Time report metrics against bad data-entry dates
# (e.g. a stray "0020" or "0226" year) that would otherwise blow the average up to
# hundreds of thousands of days. Any day-delta outside this bound is treated as bad
# data and excluded rather than allowed to skew the average.
MAX_PLAUSIBLE_DAY_DELTA = 3650  # ~10 years
from dateutil.relativedelta import relativedelta
from collections import Counter
from collections import Counter, defaultdict 
from tracker.utils import (
    get_completion_percentage, get_otif_percentage, get_overall_status,
    get_schedule_status, get_next_milestone,get_final_project_otif

)
from django.core.cache import cache
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

from django.template.loader import render_to_string
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse, QueryDict
import csv
from itertools import groupby
from operator import attrgetter
import json
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.core.mail import send_mail
from planner.models import Project as PlannerProject
from django.utils.html import escape


def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('tracker_index')
    else:
        form = AuthenticationForm()
    return render(request, 'tracker/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('tracker_login')


def signup_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "User registered successfully! Please log in.")
            return redirect('tracker_login')
    else:
        form = UserCreationForm()
    return render(request, 'tracker/signup.html', {'form': form})

@login_required
def index(request):
    show_archived = request.GET.get('archived') == '1'
    projects = Project.objects.filter(is_archived=show_archived).select_related('segment_con').prefetch_related('stages').all()
    context = {
        'projects': projects,
        'all_segments': trackerSegment.objects.all(),
        'all_team_leads': Employee.objects.filter(designation='TEAM_LEAD'),
        'show_archived': show_archived
    }
    return render(request, 'tracker/index.html', context)

@login_required
def new_project(request):
    if request.method == 'POST':
        code = request.POST['code']
        if Project.objects.filter(code=code).exists():
            messages.error(request, "Project code already exists.")
            return render(request, 'tracker/project_form.html', {
                'segments': trackerSegment.objects.all(),
                'team_leads': Employee.objects.filter(designation='TEAM_LEAD')
                })

        segment_id = request.POST.get('segment')
        segment_con = trackerSegment.objects.get(id=segment_id) if segment_id else None

        # Get the Team Lead
        team_lead_id = request.POST.get('team_lead')
        team_lead = Employee.objects.get(id=team_lead_id) if team_lead_id else None

        project = Project.objects.create(
            code=code, customer_name=request.POST['customer_name'],
            value=request.POST['value'], so_punch_date=parse_date(request.POST['so_punch_date']),
            segment_con=segment_con,
            team_lead=team_lead,
            description=request.POST.get('description', '')
        )

        # ... stage creation logic is unchanged ...
        for stage_name, _ in Stage.AUTOMATION_STAGES:
            Stage.objects.create(project=project, name=stage_name, stage_type='Automation')
        for stage_name, _ in Stage.EMULATION_STAGES:
            Stage.objects.create(project=project, name=stage_name, stage_type='Emulation')

        messages.success(request, "Project created successfully!")
        return redirect('tracker_project_detail', project_id=project.id)

    # Pass team leads to context for the GET request
    context = {
        'segments': trackerSegment.objects.all(),
        'team_leads': Employee.objects.filter(designation='TEAM_LEAD')
    }
    return render(request, 'tracker/project_form.html', context)

@login_required
def edit_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        # Update project fields from the form
        project.customer_name = request.POST['customer_name']
        project.value = request.POST['value']
        project.so_punch_date = parse_date(request.POST['so_punch_date'])
        project.description = request.POST.get('description', '')

        segment_id = request.POST.get('segment')
        project.segment_con = trackerSegment.objects.get(id=segment_id) if segment_id else None

        team_lead_id = request.POST.get('team_lead')
        project.team_lead = Employee.objects.get(id=team_lead_id) if team_lead_id else None

        project.save()
        messages.success(request, f"Project '{project.code}' updated successfully!")
        return redirect('tracker_project_detail', project_id=project.id)

    # For a GET request, show the form pre-filled with project data
    context = {
        'project': project,
        'segments': trackerSegment.objects.all(),
        'team_leads': Employee.objects.filter(designation='TEAM_LEAD')
    }
    return render(request, 'tracker/project_form.html', context)

@login_required
def project_detail(request, project_id):
    project = get_object_or_404(Project.objects.select_related('segment_con', 'team_lead'), pk=project_id)

    contact_persons = ContactPerson.objects.all()

    planner_project = PlannerProject.objects.filter(tracker_project=project).first()

    # Handle AJAX request for loading more comments
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.GET.get('action') == 'load_comments':
        offset = int(request.GET.get('offset', 0))
        limit = 5
        comments_qs = project.comments.select_related('added_by').order_by('-created_at')[offset:offset+limit]
        comments = sorted(list(comments_qs), key=lambda x: x.created_at)
        for comment in comments:
            if comment.added_by:
                comment.added_by.username = comment.added_by.get_full_name() or comment.added_by.username
        
        html = render_to_string('tracker/partials/project_comments_partial.html', {'comments': comments, 'request': request}, request=request)
        modals_html = render_to_string('tracker/partials/project_comments_modals_partial.html', {'comments': comments, 'request': request}, request=request)
        return JsonResponse({'html': html, 'modals_html': modals_html, 'has_more': project.comments.count() > (offset + limit)})

    if request.method == 'POST':
        active_tab = request.POST.get('active_tab', 'automation')

        # --- Handle Note/Remark Addition First ---
        if 'add_project_comment' in request.POST:
            note_text = request.POST.get('note_text')
            
            if note_text:
                ProjectComment.objects.create(project=project, text=note_text, added_by=request.user)
                # messages.success(request, "Note added successfully.")
            else:
                messages.error(request, "Please enter a note to save.")
            
            base_url = reverse('tracker_project_detail', args=[project.id])
            # Redirect with a hash to scroll to the notes section
            redirect_url = f'{base_url}?active_tab={active_tab}#project-notes'
            return HttpResponseRedirect(redirect_url)

        stages_to_save = []
        if 'save_all_automation' in request.POST:
            stages_to_save = project.stages.filter(stage_type='Automation')
        elif 'save_all_emulation' in request.POST:
            stages_to_save = project.stages.filter(stage_type='Emulation')
        elif 'stage_id' in request.POST:
            stage_id = request.POST.get('stage_id')
            stages_to_save = project.stages.filter(id=stage_id)

        success_message = "Changes saved successfully!"
        for stage in stages_to_save:

            # Get new values from the form

            new_planned_start_str = request.POST.get(f'planned_start_date_{stage.id}')
            new_planned_str = request.POST.get(f'planned_date_{stage.id}')

            new_status = request.POST.get(f'status_{stage.id}') or "Not started"
            actual_date_val = request.POST.get(f'actual_date_{stage.id}')
            new_completion_percentage = request.POST.get(f'completion_percentage_{stage.id}')
            


            # Safely parse date strings

            new_planned_start = parse_date(new_planned_start_str) if new_planned_start_str else None
            new_planned = parse_date(new_planned_str) if new_planned_str else None
            new_actual = parse_date(actual_date_val) if new_status == 'Completed' and actual_date_val else None
            
            # Auto-fill actual date if Completed and missing (Consistency with AJAX)
            if new_status == 'Completed' and not new_actual:
                new_actual = timezone.now().date()

            if new_status == 'Completed':
                new_completion = 100
            elif new_status == 'Not started':
                new_completion = 0
            elif new_completion_percentage is not None:
                new_completion = int(new_completion_percentage) if new_completion_percentage else 0
            else:
                new_completion = stage.completion_percentage



            # Log changes to history

            if stage.planned_start_date != new_planned_start:
                StageHistory.objects.create(stage=stage, changed_by=request.user, field_name="Planned Start Date", old_value=str(stage.planned_start_date), new_value=str(new_planned_start))
            if stage.planned_date != new_planned:
                StageHistory.objects.create(stage=stage, changed_by=request.user, field_name="Planned Finish Date", old_value=str(stage.planned_date), new_value=str(new_planned))
            if stage.status != new_status:
                StageHistory.objects.create(stage=stage, changed_by=request.user, field_name="Status", old_value=stage.status, new_value=new_status)
            if stage.actual_date != new_actual:
                StageHistory.objects.create(stage=stage, changed_by=request.user, field_name="Actual Finish Date", old_value=str(stage.actual_date), new_value=str(new_actual))
            if stage.completion_percentage != new_completion:
                StageHistory.objects.create(stage=stage, changed_by=request.user, field_name="% Completion", old_value=str(stage.completion_percentage), new_value=str(new_completion))
            
            stage.planned_start_date = new_planned_start
            stage.planned_date = new_planned
            stage.status = new_status
            stage.actual_date = new_actual
            stage.completion_percentage = new_completion
            stage.save()

        if 'save_all_automation' in request.POST:
            success_message = "Automation Stages saved successfully!"
        elif 'save_all_emulation' in request.POST:
            success_message = "Emulation Stages saved successfully!"
        elif 'stage_id' in request.POST:
            stage_name = stages_to_save.first().name if stages_to_save else ''
            success_message = f"Stage '{stage_name}' updated successfully!"

        cache.delete(f'project_detail_{project_id}')
        messages.success(request, success_message)
        base_url = reverse('tracker_project_detail', args=[project.id])
        redirect_url = f'{base_url}?active_tab={active_tab}'
        return HttpResponseRedirect(redirect_url)

    # Filter stages based on status if provided
    status_filter = request.GET.get('status_filter')

    automation_stages_qs = Stage.objects.filter(project=project, stage_type='Automation').prefetch_related('remarks', 'history')
    emulation_stages_qs = Stage.objects.filter(project=project, stage_type='Emulation').prefetch_related('remarks', 'history')
    automation_stages_qs = Stage.objects.filter(project=project, stage_type='Automation').prefetch_related('remarks__added_by', 'history__changed_by')
    emulation_stages_qs = Stage.objects.filter(project=project, stage_type='Emulation').prefetch_related('remarks__added_by', 'history__changed_by')
    
    if status_filter:
        automation_stages_qs = automation_stages_qs.filter(status=status_filter)
        emulation_stages_qs = emulation_stages_qs.filter(status=status_filter)
    
    automation_order = {name: i for i, (name, _) in enumerate(Stage.AUTOMATION_STAGES)}
    emulation_order = {name: i for i, (name, _) in enumerate(Stage.EMULATION_STAGES)}
    automation_stages = sorted(list(automation_stages_qs), key=lambda s: automation_order.get(s.name, 99))
    emulation_stages = sorted(list(emulation_stages_qs), key=lambda s: emulation_order.get(s.name, 99))

    all_stages = automation_stages + emulation_stages
    
    updates = project.updates.select_related('author', 'raised_by').prefetch_related('who_contact', 'remarks__added_by').all()[:5]
    updates_count = project.updates.count()
    open_updates_count = project.updates.filter(status__in=['Open', 'In Progress']).count()
    
    recent_activity = StageHistory.objects.select_related('stage', 'changed_by').filter(stage__project=project).order_by('-changed_at')[:5]
    last_update_obj = StageHistory.objects.filter(stage__project=project).order_by('-changed_at').first()
    last_update_time = last_update_obj.changed_at if last_update_obj else project.so_punch_date
    
    applicable_auto_stages = [s for s in automation_stages if s.status != "Not Applicable"]
    last_completed_auto_index = -1
    for i, stage in enumerate(applicable_auto_stages):
        if stage.status == "Completed": last_completed_auto_index = i
    timeline_progress_auto = 0
    total_auto_segments = len(applicable_auto_stages) - 1
    if last_completed_auto_index >= 0 and total_auto_segments > 0:
        timeline_progress_auto = round((last_completed_auto_index / total_auto_segments) * 100)

    applicable_emu_stages = [s for s in emulation_stages if s.status != "Not Applicable"]
    last_completed_emu_index = -1
    for i, stage in enumerate(applicable_emu_stages):
        if stage.status == "Completed": last_completed_emu_index = i
    timeline_progress_emu = 0
    total_emu_segments = len(applicable_emu_stages) - 1
    if last_completed_emu_index >= 0 and total_emu_segments > 0:
        timeline_progress_emu = round((last_completed_emu_index / total_emu_segments) * 100)

    total_comments_count = project.comments.count()
    initial_limit = 5
    recent_comments = project.comments.select_related('added_by').order_by('-created_at')[:initial_limit]
    project_comments = sorted(list(recent_comments), key=lambda x: x.created_at)
    for comment in project_comments:
        if comment.added_by:
            comment.added_by.username = comment.added_by.get_full_name() or comment.added_by.username
    
    context = {
        'project': project,
        'automation_stages': automation_stages,
        'emulation_stages': emulation_stages,
        'updates': updates,
        'updates_count': updates_count,
        'open_updates_count': open_updates_count,
        'completion_percentage': get_completion_percentage(all_stages),
        'timeline_progress_auto': timeline_progress_auto,
        'timeline_progress_emu': timeline_progress_emu,
        'overall_otif_percentage': get_otif_percentage(all_stages),
        'project_otif': get_final_project_otif(all_stages),
        'overall_status': get_overall_status(all_stages),
        'automation_schedule_status': get_schedule_status(automation_stages),
        'emulation_schedule_status': get_schedule_status(emulation_stages),
        'next_automation_milestone': get_next_milestone(automation_stages),
        'next_emulation_milestone': get_next_milestone(emulation_stages),
        'last_update_time': last_update_time,
        'recent_activity': recent_activity,
        'project_comments': project_comments,
        'total_comments_count': total_comments_count,
        'initial_comments_limit': initial_limit,

        'contact_persons': contact_persons,
        'status_choices': Stage.STATUS_CHOICES,
        'selected_status_filter': status_filter,
        'planner_project': planner_project,
        'update_status_choices': ProjectUpdate.STATUS_CHOICES,

    }
    
    return render(request, 'tracker/project_detail.html', context)

@login_required
def delete_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    project.delete()
    messages.success(request, "Project deleted successfully.")
    return redirect('tracker_index')

@login_required
def toggle_archive_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        project.is_archived = not project.is_archived
        project.save()
        status_str = "archived" if project.is_archived else "unarchived"
        messages.success(request, f"Project '{project.code}' has been {status_str}.")
    return redirect('tracker_project_detail', project_id=project.id)

@login_required
def edit_remark(request, remark_id):
    remark = get_object_or_404(StageRemark, pk=remark_id)
    project_id = remark.stage.project.id
    # Security check: only the author or a staff member can edit
    if request.user == remark.added_by or request.user.is_staff:
        if request.method == 'POST':
            new_text = request.POST.get('remark_text')
            if new_text:
                remark.text = new_text
                remark.save()
                messages.success(request, "Remark updated successfully.")
    else:
        messages.error(request, "You do not have permission to edit this remark.")
    return redirect('tracker_project_detail', project_id=project_id)

@login_required
def delete_remark(request, remark_id):
    remark = get_object_or_404(StageRemark, pk=remark_id)
    project_id = remark.stage.project.id
    # Security check: only the author or a staff member can delete
    if request.user == remark.added_by or request.user.is_staff:
        if request.method == 'POST':
            remark.delete()
            messages.success(request, "Remark deleted successfully.")
    else:
        messages.error(request, "You do not have permission to delete this remark.")
    return redirect('tracker_project_detail', project_id=project_id)

@login_required
def dashboard(request):
    today = timezone.now().date()
    
    # --- Dynamic Financial Year Logic ---
    # 1. Identify the current Financial Year start year
    if today.month >= 4:
        current_fy_year = today.year
    else:
        current_fy_year = today.year - 1

    # 2. Get all distinct years from project dates to build the list
    project_dates = Project.objects.filter(so_punch_date__isnull=False).values_list('so_punch_date', flat=True)
    fy_years = {current_fy_year} # Start with current FY
    for p_date in project_dates:
        if p_date.month >= 4:
            fy_years.add(p_date.year)
        else:
            fy_years.add(p_date.year - 1)
            
    # 3. Build the period map
    period_map = {}
    for year in sorted(list(fy_years), reverse=True):
        start = date(year, 4, 1)
        end = date(year + 1, 3, 31)
        label = f"FY {str(year)[-2:]}-{str(year + 1)[-2:]}"
        key = f"fy_{year}"
        period_map[key] = (label, start, end)

    default_period = f"fy_{current_fy_year}"
    period = request.GET.get('period', default_period)
    custom_start = request.GET.get('start_date_custom')
    custom_end = request.GET.get('end_date_custom')
    start_date, end_date = None, None
    display_period = "Custom"
    if custom_start and custom_end:
        start_date = parse_date(custom_start)
        end_date = parse_date(custom_end)
    else:
        if period not in period_map:
            period = default_period
        
        label, start_date, end_date = period_map[period]
        display_period = label

    completed_early_ids = Project.objects.filter(
        is_archived=False, stages__name='Handover', stages__status='Completed', stages__actual_date__lt=start_date
    ).values_list('id', flat=True)
    live_projects = Project.objects.filter(is_archived=False).filter(
        Q(so_punch_date__lte=end_date) | Q(so_punch_date__isnull=True)
    ).exclude(id__in=completed_early_ids).select_related('segment_con', 'team_lead').prefetch_related('stages').distinct()

    # --- CHRONIC PROJECTS LOGIC CORRECTION ---
    chronic_period = request.GET.get('chronic_period', '1y')
    chronic_cutoff_date = today
    if chronic_period == '6m': chronic_cutoff_date = today - relativedelta(months=6)
    elif chronic_period == '1y': chronic_cutoff_date = today - relativedelta(years=1)
    elif chronic_period == '2y': chronic_cutoff_date = today - relativedelta(years=2)

    # NEW, ROBUST QUERY
    # 1. First, get the IDs of all projects that are genuinely completed.
    completed_project_ids = Project.objects.filter(
        is_archived=False,
        stages__name='Handover',
        stages__status='Completed'
    ).values_list('id', flat=True)

    # 2. Then, find projects that are older than the cutoff AND are NOT in the completed list.
    chronic_projects = Project.objects.filter(is_archived=False).exclude(
        id__in=completed_project_ids
    ).filter(
        so_punch_date__lt=chronic_cutoff_date
    ).select_related('segment_con').order_by('so_punch_date')
    # --- END OF CORRECTION ---

    completed_stages = Stage.objects.filter(project__in=live_projects, status='Completed')
    if period != 'all' or (custom_start and custom_end):
        completed_stages = completed_stages.filter(actual_date__range=[start_date, end_date])
    total_completed_stages = completed_stages.count()
    on_time_stages = completed_stages.filter(planned_date__isnull=False, actual_date__lte=F('planned_date')).count()
    department_otif = round((on_time_stages / total_completed_stages) * 100, 1) if total_completed_stages > 0 else 0
    total_live_projects = live_projects.count()
    
    # Optimized Value Calculation
    total_live_value = live_projects.aggregate(total=Sum('value'))['total'] or 0
    
    # Optimized Loop for Status, Completion, and Counts
    status_counts = Counter()
    labels = []
    on_track_data, at_risk_data, delayed_data = [], [], []
    segment_counts = Counter()
    team_lead_counts = Counter()
    
    segment_values = defaultdict(float)
    team_lead_values = defaultdict(float)
    status_values = defaultdict(float)
    age_counts = defaultdict(int)
    
    for project in live_projects:
        # Use utils functions with the prefetched stages list to avoid N+1 queries
        stages = list(project.stages.all())
        status = get_overall_status(stages)
        completion = get_completion_percentage(stages) or 0
        
        status_counts[status] += 1
        labels.append(project.code)

        # Aggregate Segment & Team Lead Data
        seg_name = project.segment_con.name if project.segment_con else 'Unassigned'
        segment_counts[seg_name] += 1
        tl_name = project.team_lead.name if project.team_lead else 'Unassigned'
        team_lead_counts[tl_name] += 1
        
        # --- NEW: Value Aggregations ---
        val = float(project.value) if project.value else 0.0
        segment_values[seg_name] += val
        team_lead_values[tl_name] += val
        status_values[status] += val

        # --- NEW: Project Age Split ---
        if project.so_punch_date:
            age_days = (today - project.so_punch_date).days
            if age_days < 180: age_bucket = 'Below 6 Months'
            elif age_days < 365: age_bucket = '6 Months - 1 Year'
            elif age_days < 547: age_bucket = '1 - 1.5 Years'
            elif age_days < 730: age_bucket = '1.5 - 2 Years'
            else: age_bucket = 'More than 2 Years'
            age_counts[age_bucket] += 1

        
        if completion >= 80: on_track_data.append(completion); at_risk_data.append(0); delayed_data.append(0)
        elif completion >= 40: on_track_data.append(0); at_risk_data.append(completion); delayed_data.append(0)
        else: on_track_data.append(0); at_risk_data.append(0); delayed_data.append(completion)
    
    # Derive counts from the consistent status calculation
    active_live_projects = status_counts['In Progress'] + status_counts['Not started']
    delayed_live_projects = status_counts['Hold']
    
    # --- FIX: Enforce consistent order for Status Chart Colors ---
    ordered_statuses = ['Completed', 'In Progress', 'Hold', 'Not started']
    color_map = {
        'Completed': 'rgba(40, 167, 69, 0.7)',   # Green
        'In Progress': 'rgba(255, 193, 7, 0.7)', # Yellow
        'Hold': 'rgba(220, 53, 69, 0.7)',        # Red
        'Not started': 'rgba(108, 117, 125, 0.7)', # Grey
        'Not Applicable': 'rgba(200, 200, 200, 0.5)'
    }
    
    final_status_labels = []
    final_status_data = []
    final_status_colors = []
    
    for s in ordered_statuses + [k for k in status_counts.keys() if k not in ordered_statuses]:
        if status_counts[s] > 0:
            final_status_labels.append(s)
            final_status_data.append(status_counts[s])
            final_status_colors.append(color_map.get(s, 'rgba(100, 100, 100, 0.7)'))
            
    # --- NEW: Prepare Status Value Data (Consistent Coloring) ---
    final_status_value_labels = []
    final_status_value_data = []
    final_status_value_colors = []
    for s in ordered_statuses + [k for k in status_values.keys() if k not in ordered_statuses]:
        if status_values[s] > 0:
            final_status_value_labels.append(s)
            final_status_value_data.append(status_values[s])
            final_status_value_colors.append(color_map.get(s, 'rgba(100, 100, 100, 0.7)'))

    # --- NEW: Prepare Age Data (Ordered) ---
    age_buckets_order = ['Below 6 Months', '6 Months - 1 Year', '1 - 1.5 Years', '1.5 - 2 Years', 'More than 2 Years']
    age_labels = []
    age_data = []
    for bucket in age_buckets_order:
        if age_counts[bucket] > 0:
            age_labels.append(bucket)
            age_data.append(age_counts[bucket])


    context = {
        'total_projects': total_live_projects, 'active_projects': active_live_projects, 'delayed_projects': delayed_live_projects, 'total_value': total_live_value,
        'department_otif': department_otif, 
        'recent_projects': Project.objects.select_related('segment_con').prefetch_related('stages').order_by('-so_punch_date')[:5],
        'labels': labels, 'on_track_data': on_track_data, 'at_risk_data': at_risk_data, 'delayed_data': delayed_data,
        'status_labels': final_status_labels, 'status_data': final_status_data, 'status_colors': final_status_colors,
        'segment_labels': list(segment_counts.keys()), 'segment_data': list(segment_counts.values()),
        'team_lead_labels': list(team_lead_counts.keys()), 'team_lead_data': list(team_lead_counts.values()),
        'segment_value_labels': list(segment_values.keys()), 'segment_value_data': list(segment_values.values()),
        'team_lead_value_labels': list(team_lead_values.keys()), 'team_lead_value_data': list(team_lead_values.values()),
        'status_value_labels': final_status_value_labels, 'status_value_data': final_status_value_data, 'status_value_colors': final_status_value_colors,
        'age_labels': age_labels, 'age_data': age_data,
        'selected_period_display': display_period,
        'custom_start_date': custom_start, 'custom_end_date': custom_end, 'chronic_projects': chronic_projects, 'selected_chronic_period': chronic_period,
        'period_filter_options': period_map, 'selected_period_key': period,
    }
    return render(request, 'tracker/dashboard.html', context)

# tracker/views.py

@login_required
def project_reports(request):
    # --- Session logic for retaining filters ---
    
    # 1. Handle explicit reset action
    if 'reset' in request.GET:
        if 'report_filters' in request.session:
            del request.session['report_filters']
        return redirect('project_reports')

    # Determine which query parameters to use: from request or from session
    if not request.GET: # If no filters in URL
        # Check if filters are stored in session
        if 'report_filters' in request.session:
            # Rebuild query string and redirect
            saved_filters = request.session.get('report_filters', {})
            if saved_filters:
                query_dict = QueryDict(mutable=True)
                query_dict.update(saved_filters)
                return redirect(f"{reverse('project_reports')}?{query_dict.urlencode()}")
    else:
        # Filters are in the URL, save them to the session
        request.session['report_filters'] = request.GET.dict()

    # The rest of the view logic now uses the active query parameters
    query_params = request.GET or request.session.get('report_filters', {})
    
    # --- The FIX is in this line: We add select_related and prefetch_related ---
    projects_qs = Project.objects.filter(is_archived=False).select_related('segment_con', 'team_lead').prefetch_related('stages').all()

    # --- Dynamic Financial Year Logic (Same as Dashboard) ---
    today = timezone.now().date()
    # A handful of seed/legacy records have corrupted planned_date values (e.g. year
    # 0002, 0020, 0226 — clearly truncated typos). Without a floor, "All Time" bucketing
    # walks the trend chart's x-axis back to the year those records claim, and they also
    # inflate "planned" counts in every period. Treat anything older than this as bad data.
    plausible_planned_date_floor = today - timedelta(days=MAX_PLAUSIBLE_DAY_DELTA)
    if today.month >= 4:
        current_fy_year = today.year
    else:
        current_fy_year = today.year - 1

    # Get all distinct years from project dates
    project_dates = Project.objects.filter(so_punch_date__isnull=False).values_list('so_punch_date', flat=True)
    fy_years = {current_fy_year}
    for p_date in project_dates:
        if p_date.month >= 4:
            fy_years.add(p_date.year)
        else:
            fy_years.add(p_date.year - 1)
            
    fy_options = []
    for year in sorted(list(fy_years), reverse=True):
        label = f"FY {str(year)[-2:]}-{str(year + 1)[-2:]}"
        value = str(year)
        fy_options.append((value, label))

    # --- Get standard filter values ---
    # getlist needs a QueryDict, not a regular dict
    params_for_getlist = QueryDict(mutable=True)
    params_for_getlist.update(query_params)
    selected_segment_ids = params_for_getlist.getlist('segments')
    selected_team_lead_ids = params_for_getlist.getlist('team_leads')
    selected_stage_keys = params_for_getlist.getlist('stages')

    min_value = query_params.get('min_value')
    max_value = query_params.get('max_value')
    selected_fy = query_params.get('financial_year')

    actual_start_date = parse_date(query_params.get('actual_start_date')) if query_params.get('actual_start_date') else None
    actual_end_date = parse_date(query_params.get('actual_end_date')) if query_params.get('actual_end_date') else None

    # --- Resolve the reporting period (Planned Date Range) ---
    # Defaults to the current month so the page opens on a live, meaningful window
    # instead of silently dumping "all time". "period" drives this; explicit custom
    # dates imply period=custom even if the param itself wasn't sent.
    period = query_params.get('period')
    if not period:
        period = 'custom' if (query_params.get('planned_start_date') and query_params.get('planned_end_date')) else 'this_month'

    def _month_range(d):
        month_start = d.replace(day=1)
        month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
        return month_start, month_end

    def _quarter_range(d):
        q_start_month = ((d.month - 1) // 3) * 3 + 1
        q_start = date(d.year, q_start_month, 1)
        q_end = (q_start + relativedelta(months=3)) - timedelta(days=1)
        return q_start, q_end

    if period == 'this_quarter':
        start_date, end_date = _quarter_range(today)
    elif period == 'all':
        start_date, end_date = date(2000, 1, 1), today
    elif period == 'custom':
        start_date = parse_date(query_params.get('planned_start_date')) if query_params.get('planned_start_date') else None
        end_date = parse_date(query_params.get('planned_end_date')) if query_params.get('planned_end_date') else None
        if not (start_date and end_date):
            period = 'this_month'
            start_date, end_date = _month_range(today)
    else:  # 'this_month' (default)
        period = 'this_month'
        start_date, end_date = _month_range(today)

    # Financial Year overrides the resolved period entirely (same precedence as before)
    if selected_fy:
        year = int(selected_fy)
        fy_start = date(year, 4, 1)
        fy_end = date(year + 1, 3, 31)
        start_date = fy_start
        end_date = fy_end

        # Match Dashboard Logic: Live Projects in Period (Carry-over + New)
        # 1. Exclude projects completed before the start of the FY
        completed_early_ids = Project.objects.filter(
            is_archived=False,
            stages__name='Handover',
            stages__status='Completed',
            stages__actual_date__lt=fy_start
        ).values_list('id', flat=True)

        # 2. Include projects punched on/before end of period (or null)
        projects_qs = projects_qs.filter(
            Q(so_punch_date__lte=fy_end) | Q(so_punch_date__isnull=True)
        ).exclude(id__in=completed_early_ids)

        if selected_stage_keys:
            projects_qs = projects_qs.filter(stages__name__in=selected_stage_keys)
    else:
        # Scope to projects with a stage (optionally restricted to selected stage types)
        # due in this window, including backlog carried forward from earlier windows.
        planned_window_stages = Stage.objects.filter(
            planned_date__isnull=False,
            planned_date__gte=plausible_planned_date_floor,
            planned_date__lte=end_date,
        ).exclude(Q(status='Not Applicable') | (Q(status='Completed') & Q(actual_date__lt=start_date)))
        if selected_stage_keys:
            planned_window_stages = planned_window_stages.filter(name__in=selected_stage_keys)
        projects_qs = projects_qs.filter(id__in=planned_window_stages.values_list('project_id', flat=True))

    # Actual Date Range: independent filter, applies on top of the above either way
    if actual_start_date and actual_end_date:
        actual_window_stages = Stage.objects.filter(
            status='Completed',
            actual_date__range=[actual_start_date, actual_end_date],
        )
        if selected_stage_keys:
            actual_window_stages = actual_window_stages.filter(name__in=selected_stage_keys)
        projects_qs = projects_qs.filter(id__in=actual_window_stages.values_list('project_id', flat=True))

    has_explicit_period = period != 'all'

    if selected_segment_ids:
        projects_qs = projects_qs.filter(segment_con__id__in=selected_segment_ids)

    if selected_team_lead_ids:
        projects_qs = projects_qs.filter(team_lead__id__in=selected_team_lead_ids)

    if min_value:
        try:
            projects_qs = projects_qs.filter(value__gte=float(min_value))
        except (ValueError, TypeError): pass
    if max_value:
        try:
            projects_qs = projects_qs.filter(value__lte=float(max_value))
        except (ValueError, TypeError): pass

    # Which stages the rest of the report (summary table, charts, cross-tab) covers
    if selected_stage_keys:
        stage_names_to_report = [(k, v) for k, v in Stage.STAGE_NAMES if k in selected_stage_keys]
    else:
        stage_names_to_report = Stage.STAGE_NAMES
    stage_keys_to_report = [k for k, _ in stage_names_to_report]

    # --- Check for the 'hide_completed' filter ---
    hide_completed = query_params.get('hide_completed') == '1'
    if hide_completed:
        completed_project_ids = Project.objects.filter(
            is_archived=False,
            stages__name='Handover',
            stages__status='Completed'
        ).values_list('id', flat=True)
        projects_qs = projects_qs.exclude(id__in=completed_project_ids)

    # --- Active filter chips (for quick removal in the UI) ---
    active_filters = []
    base_qd = request.GET.copy()

    if selected_fy:
        fy_label = dict(fy_options).get(selected_fy, selected_fy)
        qd = base_qd.copy(); qd.pop('financial_year', None)
        active_filters.append({'label': f'FY: {fy_label}', 'href': f"?{qd.urlencode()}"})

    if selected_segment_ids:
        names = list(trackerSegment.objects.filter(id__in=selected_segment_ids).values_list('name', flat=True))
        qd = base_qd.copy(); qd.pop('segments', None)
        active_filters.append({'label': f"Segments: {', '.join(names)}", 'href': f"?{qd.urlencode()}"})

    if selected_team_lead_ids:
        names = list(Employee.objects.filter(id__in=selected_team_lead_ids).values_list('name', flat=True))
        qd = base_qd.copy(); qd.pop('team_leads', None)
        active_filters.append({'label': f"Team Leads: {', '.join(names)}", 'href': f"?{qd.urlencode()}"})

    if not selected_fy and (query_params.get('period') or query_params.get('planned_start_date')):
        qd = base_qd.copy(); qd.pop('period', None); qd.pop('planned_start_date', None); qd.pop('planned_end_date', None)
        period_labels = {'this_month': 'This Month', 'this_quarter': 'This Quarter', 'all': 'All Time'}
        period_label = period_labels.get(period, f"{start_date} to {end_date}")
        active_filters.append({'label': f"Planned Period: {period_label}", 'href': f"?{qd.urlencode()}"})

    if selected_stage_keys:
        stage_display_map = dict(Stage.STAGE_NAMES)
        names = [stage_display_map.get(k, k) for k in selected_stage_keys]
        qd = base_qd.copy(); qd.pop('stages', None)
        active_filters.append({'label': f"Stages: {', '.join(names)}", 'href': f"?{qd.urlencode()}"})

    if actual_start_date and actual_end_date:
        qd = base_qd.copy(); qd.pop('actual_start_date', None); qd.pop('actual_end_date', None)
        active_filters.append({'label': f"Actual Completed: {actual_start_date} to {actual_end_date}", 'href': f"?{qd.urlencode()}"})

    if min_value or max_value:
        qd = base_qd.copy(); qd.pop('min_value', None); qd.pop('max_value', None)
        active_filters.append({'label': f"Value: {min_value or '0'} - {max_value or '∞'} Cr", 'href': f"?{qd.urlencode()}"})

    if hide_completed:
        qd = base_qd.copy(); qd.pop('hide_completed', None)
        active_filters.append({'label': 'Hide Completed', 'href': f"?{qd.urlencode()}"})

    # --- Capture QS for Charts (Before Hide Completed) ---
    chart_projects_qs = projects_qs

    # --- Chart Click Filtering ---
    # The trend chart is an aggregate across all currently-reported stages (no per-stage
    # picker), so a bar click scopes to stage_keys_to_report rather than one stage name.
    chart_filter_month = query_params.get('chart_filter_month')
    chart_filter_type = query_params.get('chart_filter_type')

    if chart_filter_month and chart_filter_type:
        try:
            filter_date = parse_date(chart_filter_month)
            if filter_date:
                if chart_filter_type == 'planned':
                    # Matches the trend chart's backlog-inclusive "planned" bucket: due by
                    # this month's end and not already resolved before this month started.
                    month_start = filter_date.replace(day=1)
                    month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
                    matching_stage_ids = Stage.objects.filter(
                        name__in=stage_keys_to_report,
                        planned_date__gte=plausible_planned_date_floor,
                        planned_date__lte=month_end,
                    ).exclude(
                        Q(status='Not Applicable') | (Q(status='Completed') & Q(actual_date__lt=month_start))
                    ).values_list('project_id', flat=True)
                    projects_qs = projects_qs.filter(id__in=matching_stage_ids)
                elif chart_filter_type == 'actual':
                    matching_stage_ids = Stage.objects.filter(
                        name__in=stage_keys_to_report,
                        actual_date__year=filter_date.year,
                        actual_date__month=filter_date.month
                    ).values_list('project_id', flat=True)
                    projects_qs = projects_qs.filter(id__in=matching_stage_ids)
        except (ValueError, TypeError):
            pass

    distinct_projects = projects_qs.distinct()
    distinct_chart_projects = chart_projects_qs.distinct()
    chart_project_ids = chart_projects_qs.values_list('id', flat=True).distinct()

    # --- Prepare projects with their detailed summaries ---
    projects_with_details = []
    automation_order = {name: i for i, (name, _) in enumerate(Stage.AUTOMATION_STAGES)}
    emulation_order = {name: i for i, (name, _) in enumerate(Stage.EMULATION_STAGES)}
    for project in distinct_projects:
        all_stages = list(project.stages.all())
        auto_stages = sorted([s for s in all_stages if s.stage_type == 'Automation'], key=lambda s: automation_order.get(s.name, 99))
        emu_stages = sorted([s for s in all_stages if s.stage_type == 'Emulation'], key=lambda s: emulation_order.get(s.name, 99))
        projects_with_details.append({
            'project': project,
            'otif': project.get_otif_percentage(),
            'next_auto_milestone': get_next_milestone(auto_stages),
            'next_emu_milestone': get_next_milestone(emu_stages),
            'auto_schedule': get_schedule_status(auto_stages),
            'emu_schedule': get_schedule_status(emu_stages),
        })

    # --- NEW: Stage Bottleneck Analysis (Top Delayed Stages) ---
    # Count stages where Actual > Planned OR (Status is active AND Today > Planned)
    today = timezone.now().date()
    delayed_stages_qs = Stage.objects.filter(
        project__in=distinct_projects,
        planned_date__gte=plausible_planned_date_floor,
    ).exclude(
        status='Not Applicable'
    ).filter(
        Q(actual_date__gt=F('planned_date')) |
        Q(status__in=['Not started', 'In Progress'], planned_date__lt=today)
    )
    if selected_stage_keys:
        delayed_stages_qs = delayed_stages_qs.filter(name__in=stage_keys_to_report)
    delayed_stages_qs = delayed_stages_qs.values('name').annotate(count=Count('id')).order_by('-count')

    stage_delay_labels = [item['name'] for item in delayed_stages_qs[:10]] # Top 10 bottlenecks
    stage_delay_data = [item['count'] for item in delayed_stages_qs[:10]]

    # --- NEW: Monthly Planned vs Actual Trend (aggregate across all reported stages, with
    # backlog carry-forward) ---
    # "Planned" per month is a point-in-time snapshot (like the summary table): anything
    # due by that month's end that wasn't already completed before that month started, so
    # a stage planned in an earlier month keeps showing as planned until it's resolved
    # instead of only appearing once in its origin month. Combines every stage currently
    # in scope (all stages, or just the ones picked in the "Stages" filter) into one trend.
    trend_source_qs = Stage.objects.filter(
        project_id__in=chart_project_ids,
        planned_date__isnull=False,
        planned_date__gte=plausible_planned_date_floor,
        planned_date__lte=end_date,
    ).exclude(status='Not Applicable')
    if selected_stage_keys:
        trend_source_qs = trend_source_qs.filter(name__in=stage_keys_to_report)
    trend_rows = list(trend_source_qs.values('planned_date', 'actual_date', 'status'))

    if has_explicit_period:
        trend_range_start = start_date
    else:
        earliest_planned = min((r['planned_date'] for r in trend_rows), default=end_date)
        trend_range_start = earliest_planned

    months_list = []
    cursor = trend_range_start.replace(day=1)
    end_cursor = end_date.replace(day=1)
    while cursor <= end_cursor:
        months_list.append(cursor)
        cursor = cursor + relativedelta(months=1)

    trend_planned = []
    trend_actual = []
    for month_start in months_list:
        month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
        trend_planned.append(sum(
            1 for r in trend_rows
            if r['planned_date'] <= month_end and not (
                r['status'] == 'Completed' and r['actual_date'] and r['actual_date'] < month_start
            )
        ))
        trend_actual.append(sum(
            1 for r in trend_rows
            if r['actual_date'] and month_start <= r['actual_date'] <= month_end
        ))

    trend_labels = [m.strftime('%b %Y') for m in months_list]
    trend_years = [m.year for m in months_list]
    trend_months = [m.month for m in months_list]
    trend_financial_years = []
    for m in months_list:
        if m.month >= 4:
            fy_str = f"FY {str(m.year)[-2:]}-{str(m.year + 1)[-2:]}"
        else:
            fy_str = f"FY {str(m.year - 1)[-2:]}-{str(m.year)[-2:]}"
        trend_financial_years.append(fy_str)

    stage_trend_data = {
        'labels': trend_labels,
        'years': trend_years,
        'financial_years': trend_financial_years,
        'months': trend_months,
        'planned': trend_planned,
        'actual': trend_actual,
    }

    # --- NEW: OTIF Trend (aggregate across all reported stages) ---
    otif_qs = Stage.objects.filter(
        project_id__in=chart_project_ids,
        planned_date__isnull=False,
        planned_date__gte=plausible_planned_date_floor,
    ).exclude(status='Not Applicable')
    if selected_stage_keys:
        otif_qs = otif_qs.filter(name__in=stage_keys_to_report)
    if has_explicit_period:
        otif_qs = otif_qs.filter(planned_date__range=[start_date, end_date])

    otif_qs = otif_qs.annotate(
        month=TruncMonth('planned_date')
    ).values('month').annotate(
        total=Count('id'),
        on_time=Count('id', filter=Q(actual_date__isnull=False) & Q(actual_date__lte=F('planned_date')))
    ).order_by('month')

    temp_otif = {}
    for item in otif_qs:
        if item['month']:
            temp_otif[item['month']] = {'total': item['total'], 'on_time': item['on_time']}

    otif_sorted_months = sorted(temp_otif.keys())
    otif_financial_years = []
    for m in otif_sorted_months:
        if m.month >= 4:
            fy_str = f"FY {str(m.year)[-2:]}-{str(m.year + 1)[-2:]}"
        else:
            fy_str = f"FY {str(m.year - 1)[-2:]}-{str(m.year)[-2:]}"
        otif_financial_years.append(fy_str)

    stage_otif_data = {
        'labels': [m.strftime('%b %Y') for m in otif_sorted_months],
        'years': [m.year for m in otif_sorted_months],
        'financial_years': otif_financial_years,
        'months': [m.month for m in otif_sorted_months],
        'total': [temp_otif[m]['total'] for m in otif_sorted_months],
        'on_time': [temp_otif[m]['on_time'] for m in otif_sorted_months],
    }

    # --- NEW: Per-Stage Planned vs Actual Summary (with backlog carry-forward) ---
    # "Planned" for the selected period = anything due on/before the period end that
    # wasn't already completed before the period started (so overdue/pending work from
    # earlier periods carries forward into the current one instead of disappearing).
    summary_start = start_date
    summary_end = end_date

    # Equal-length preceding window, used for period-over-period OTIF comparison.
    # Not meaningful for "All Time", which has no natural "previous period".
    if has_explicit_period:
        period_length_days = (summary_end - summary_start).days + 1
        prev_period_end = summary_start - timedelta(days=1)
        prev_period_start = prev_period_end - timedelta(days=period_length_days - 1)
    else:
        prev_period_start = prev_period_end = None

    automation_stage_keys = dict(Stage.AUTOMATION_STAGES)
    stage_summary = []
    aggregate_planned_list = []
    aggregate_actual_list = []

    total_planned = total_actual = total_pending = total_delayed = total_on_time = 0

    for stage_key, stage_display in stage_names_to_report:
        planned_backlog_qs = Stage.objects.filter(
            project_id__in=chart_project_ids,
            name=stage_key,
            planned_date__isnull=False,
            planned_date__gte=plausible_planned_date_floor,
            planned_date__lte=summary_end,
        ).exclude(
            Q(status='Not Applicable') | (Q(status='Completed') & Q(actual_date__lt=summary_start))
        ).select_related('project').order_by('planned_date')

        actual_period_qs = Stage.objects.filter(
            project_id__in=chart_project_ids,
            name=stage_key,
            status='Completed',
            actual_date__range=[summary_start, summary_end]
        ).select_related('project').order_by('actual_date')

        delayed_qs = planned_backlog_qs.filter(
            status__in=['Not started', 'In Progress'],
            planned_date__lt=today
        )

        planned_count = planned_backlog_qs.count()
        pending_count = planned_backlog_qs.exclude(status='Completed').count()
        delayed_count = delayed_qs.count()

        actual_count = actual_period_qs.count()
        on_time_count = actual_period_qs.filter(actual_date__lte=F('planned_date')).count()
        otif_pct = round((on_time_count / actual_count) * 100, 1) if actual_count else None

        total_planned += planned_count
        total_actual += actual_count
        total_pending += pending_count
        total_delayed += delayed_count
        total_on_time += on_time_count

        # Average staleness of the currently-overdue backlog, in days
        avg_delay_days = None
        overdue_planned_dates = list(delayed_qs.values_list('planned_date', flat=True))
        overdue_deltas = [(today - d).days for d in overdue_planned_dates if 0 <= (today - d).days <= MAX_PLAUSIBLE_DAY_DELTA]
        if overdue_deltas:
            avg_delay_days = round(sum(overdue_deltas) / len(overdue_deltas), 1)

        # Cycle time: planned start -> actual completion, for stages completed in this period
        avg_cycle_time = None
        cycle_pairs = list(
            actual_period_qs.exclude(planned_start_date__isnull=True).values_list('planned_start_date', 'actual_date')
        )
        cycle_deltas = [(a - p).days for p, a in cycle_pairs if 0 <= (a - p).days <= MAX_PLAUSIBLE_DAY_DELTA]
        if cycle_deltas:
            avg_cycle_time = round(sum(cycle_deltas) / len(cycle_deltas), 1)

        # Period-over-period OTIF trend
        otif_trend = None
        if has_explicit_period:
            prev_actual_qs = Stage.objects.filter(
                project_id__in=chart_project_ids,
                name=stage_key,
                status='Completed',
                actual_date__range=[prev_period_start, prev_period_end]
            )
            prev_actual_count = prev_actual_qs.count()
            if prev_actual_count and actual_count:
                prev_on_time = prev_actual_qs.filter(actual_date__lte=F('planned_date')).count()
                prev_otif = (prev_on_time / prev_actual_count) * 100
                if otif_pct > prev_otif:
                    otif_trend = 'up'
                elif otif_pct < prev_otif:
                    otif_trend = 'down'
                else:
                    otif_trend = 'flat'

        # Risk color-coding based on how much of the planned backlog is currently delayed
        if planned_count == 0:
            risk = 'none'
        else:
            delay_ratio = delayed_count / planned_count
            risk = 'high' if delay_ratio >= 0.3 else ('medium' if delay_ratio >= 0.1 else 'low')

        stage_summary.append({
            'key': stage_key,
            'display': stage_display,
            'type': 'Automation' if stage_key in automation_stage_keys else 'Emulation',
            'planned': planned_count,
            'actual': actual_count,
            'pending': pending_count,
            'delayed': delayed_count,
            'otif': otif_pct,
            'otif_trend': otif_trend,
            'avg_delay_days': avg_delay_days,
            'avg_cycle_time': avg_cycle_time,
            'risk': risk,
        })

        aggregate_planned_list.extend(
            {
                'code': s.project.code,
                'customer': s.project.customer_name,
                'stage': stage_display,
                'planned_date_raw': s.planned_date,
                'planned_date': s.planned_date.strftime('%d %b %Y') if s.planned_date else '',
                'status': s.status,
                'is_overdue': s.status in ['Not started', 'In Progress'] and s.planned_date and s.planned_date < today,
            }
            for s in planned_backlog_qs.exclude(status='Completed')
        )
        aggregate_actual_list.extend(
            {
                'code': s.project.code,
                'customer': s.project.customer_name,
                'stage': stage_display,
                'actual_date_raw': s.actual_date,
                'actual_date': s.actual_date.strftime('%d %b %Y') if s.actual_date else '',
                'planned_date': s.planned_date.strftime('%d %b %Y') if s.planned_date else '',
                'on_time': bool(s.actual_date and s.planned_date and s.actual_date <= s.planned_date),
            }
            for s in actual_period_qs
        )

    # Built stage-by-stage above; sort chronologically so the drill-down lists read
    # naturally, then drop the raw date keys used only for sorting.
    aggregate_planned_list.sort(key=lambda p: p['planned_date_raw'])
    aggregate_actual_list.sort(key=lambda a: a['actual_date_raw'])
    for p in aggregate_planned_list:
        del p['planned_date_raw']
    for a in aggregate_actual_list:
        del a['actual_date_raw']

    # --- NEW: Report-wide KPI summary (all stages combined, for the filtered set) ---
    overall_otif = round((total_on_time / total_actual) * 100, 1) if total_actual else None
    overall_otif_trend = None
    if has_explicit_period and total_actual:
        prev_overall_qs = Stage.objects.filter(
            project_id__in=chart_project_ids,
            status='Completed',
            actual_date__range=[prev_period_start, prev_period_end]
        )
        prev_overall_actual = prev_overall_qs.count()
        if prev_overall_actual:
            prev_overall_on_time = prev_overall_qs.filter(actual_date__lte=F('planned_date')).count()
            prev_overall_otif = (prev_overall_on_time / prev_overall_actual) * 100
            if overall_otif > prev_overall_otif:
                overall_otif_trend = 'up'
            elif overall_otif < prev_overall_otif:
                overall_otif_trend = 'down'
            else:
                overall_otif_trend = 'flat'

    report_kpis = {
        'total_planned': total_planned,
        'total_actual': total_actual,
        'total_pending': total_pending,
        'total_delayed': total_delayed,
        'overall_otif': overall_otif,
        'overall_otif_trend': overall_otif_trend,
    }

    # --- NEW: Who owns today's delays — cross-tab by Team Lead and Segment ---
    delayed_now_qs = Stage.objects.filter(
        project_id__in=chart_project_ids,
        planned_date__isnull=False,
        planned_date__gte=plausible_planned_date_floor,
        planned_date__lte=summary_end,
        status__in=['Not started', 'In Progress'],
        planned_date__lt=today,
    )
    if selected_stage_keys:
        delayed_now_qs = delayed_now_qs.filter(name__in=stage_keys_to_report)
    delayed_now_qs = delayed_now_qs.select_related('project__team_lead', 'project__segment_con')

    delay_by_team_lead = Counter()
    delay_by_segment = Counter()
    for s in delayed_now_qs:
        delay_by_team_lead[s.project.team_lead.name if s.project.team_lead else 'Unassigned'] += 1
        delay_by_segment[s.project.segment_con.name if s.project.segment_con else 'Unassigned'] += 1

    delay_by_team_lead_top = delay_by_team_lead.most_common()
    delay_by_segment_top = delay_by_segment.most_common()
    delay_by_team_lead_labels = [x[0] for x in delay_by_team_lead_top]
    delay_by_team_lead_data = [x[1] for x in delay_by_team_lead_top]
    delay_by_segment_labels = [x[0] for x in delay_by_segment_top]
    delay_by_segment_data = [x[1] for x in delay_by_segment_top]

    # --- NEW: Emulation Timing Analysis Trends ---
    # Categories:
    # 1. Before Dispatch
    # 2. After Dispatch but Before Go Live
    # 3. After Go Live
    emu_trend_data = defaultdict(lambda: {'cat1': 0, 'cat2': 0, 'cat3': 0})
    
    for p in distinct_chart_projects:
        stages_map = {s.name: s for s in p.stages.all()}
        
        emu = stages_map.get('Emulation Testing')
        dispatch = stages_map.get('Dispatch')
        comm = stages_map.get('Commissioning')
        
        if emu and emu.actual_date:
            month_key = emu.actual_date.replace(day=1)
            
            # Filter by date range if selected
            if has_explicit_period and not (start_date <= emu.actual_date <= end_date):
                continue
            
            dispatch_date = dispatch.actual_date if (dispatch and dispatch.actual_date) else None
            # Using planned_start_date as the 'Start Date' for Go Live (Commissioning) as per user request
            comm_start_date = comm.planned_start_date if (comm and comm.planned_start_date) else None
            
            # 1. Before Dispatch (or Dispatch not yet done)
            if not dispatch_date or emu.actual_date <= dispatch_date:
                emu_trend_data[month_key]['cat1'] += 1
            
            # 2. After Dispatch but Before Go Live (or Go Live not yet done)
            elif not comm_start_date or emu.actual_date <= comm_start_date:
                emu_trend_data[month_key]['cat2'] += 1
            
            # 3. After Go Live
            else:
                emu_trend_data[month_key]['cat3'] += 1

    sorted_months = sorted(emu_trend_data.keys())
    emu_chart_data = {
        'labels': [m.strftime('%b %Y') for m in sorted_months],
        'years': [m.year for m in sorted_months],
        'months': [m.month for m in sorted_months],
        'financial_years': [],
        'cat1': [emu_trend_data[m]['cat1'] for m in sorted_months],
        'cat2': [emu_trend_data[m]['cat2'] for m in sorted_months],
        'cat3': [emu_trend_data[m]['cat3'] for m in sorted_months],
    }
    
    for m in sorted_months:
        if m.month >= 4:
            fy_str = f"FY {str(m.year)[-2:]}-{str(m.year + 1)[-2:]}"
        else:
            fy_str = f"FY {str(m.year - 1)[-2:]}-{str(m.year)[-2:]}"
        emu_chart_data['financial_years'].append(fy_str)

    context = {

        'projects_with_details': projects_with_details,
        'all_segments': trackerSegment.objects.all(),
        'all_team_leads': Employee.objects.filter(designation='TEAM_LEAD'),
        'selected_segment_ids': [int(i) for i in selected_segment_ids],
        'selected_team_lead_ids': [int(i) for i in selected_team_lead_ids],
        'start_date': start_date, 'end_date': end_date,
        'period': period,
        'actual_start_date': actual_start_date, 'actual_end_date': actual_end_date,
        'min_value': min_value, 'max_value': max_value,
        'stage_delay_labels': stage_delay_labels, 'stage_delay_data': stage_delay_data,
        'stage_names': Stage.STAGE_NAMES, 'status_choices': Stage.STATUS_CHOICES,
        'all_automation_stage_names': Stage.AUTOMATION_STAGES,
        'all_emulation_stage_names': Stage.EMULATION_STAGES,
        'selected_stage_keys': selected_stage_keys,
        'hide_completed_active': hide_completed,
        'stage_trend_data': json.dumps(stage_trend_data),
        'stage_otif_data': json.dumps(stage_otif_data),
        'emu_timing_data': json.dumps(emu_chart_data),
        'fy_options': fy_options,
        'selected_fy': selected_fy,
        'stage_summary': stage_summary,
        'aggregate_project_lists': json.dumps({'planned': aggregate_planned_list, 'actual': aggregate_actual_list}),
        'summary_start': summary_start,
        'summary_end': summary_end,
        'active_filters': active_filters,
        'report_kpis': report_kpis,
        'delay_by_team_lead_labels': delay_by_team_lead_labels,
        'delay_by_team_lead_data': delay_by_team_lead_data,
        'delay_by_segment_labels': delay_by_segment_labels,
        'delay_by_segment_data': delay_by_segment_data,
        'saved_report_presets': SavedReportFilter.objects.filter(user=request.user) if request.user.is_authenticated else [],
        'current_query_string': request.GET.urlencode(),
    }
    return render(request, 'tracker/project_report.html', context)


@login_required
def save_report_preset(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        query_string = request.POST.get('query_string', '')
        if name:
            SavedReportFilter.objects.update_or_create(
                user=request.user, name=name,
                defaults={'query_string': query_string}
            )
            return redirect(f"{reverse('project_reports')}?{query_string}")
    return redirect('project_reports')


@login_required
def delete_report_preset(request, preset_id):
    if request.method == 'POST':
        SavedReportFilter.objects.filter(user=request.user, id=preset_id).delete()
    return redirect('project_reports')


@login_required
def project_activity(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    history_logs = StageHistory.objects.select_related('stage', 'changed_by').filter(stage__project=project).order_by('-changed_at')
    return render(request, 'tracker/project_activity.html', {
        'project': project,
        'history_logs': history_logs,
    })


from itertools import groupby
from operator import attrgetter

@login_required
def upcoming_milestones(request):
    filter_type = request.GET.get('filter', 'all')
    today = timezone.now().date()
    
    # Use the existing helper function to get the initial filtered list of stages
    stages = get_filtered_stages(filter_type)
    
    # Add select_related for performance and order by project for grouping
    stages = stages.select_related('project').order_by('project__code', 'planned_date')

    # Group the stages by project
    stages_list = list(stages)
    grouped_stages = []
    for project, group in groupby(stages_list, key=attrgetter('project')):
        grouped_stages.append({
            'project': project,
            'stages': list(group)
        })

    return render(request, 'tracker/upcoming_milestones.html', {
        'grouped_stages': grouped_stages,
        'filter_type': filter_type,
        'filter_options': [
            ('All', 'all'), ('Overdue', 'overdue'), ('Today', 'today'),
            ('Tomorrow', 'tomorrow'), ('This Week', 'this_week'),
            ('Next Week', 'next_week'), ('This Month', 'this_month'),
            ('Next Month', 'next_month'),
        ],
        'today': today,
    })


def get_filtered_stages(filter_type):
    today = date.today()
    tomorrow = today + timedelta(days=1)
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_of_next_week = end_of_week + timedelta(days=1)
    end_of_next_week = start_of_next_week + timedelta(days=6)
    start_of_month = today.replace(day=1)
    start_of_next_month = (start_of_month + relativedelta(months=1)).replace(day=1)
    end_of_month = start_of_next_month - timedelta(days=1)
    end_of_next_month = (start_of_next_month + relativedelta(months=1)) - timedelta(days=1)

    date_ranges = {
        "today": (today, today),
        "tomorrow": (tomorrow, tomorrow),
        "this_week": (start_of_week, end_of_week),
        "next_week": (start_of_next_week, end_of_next_week),
        "this_month": (start_of_month, end_of_month),
        "next_month": (start_of_next_month, end_of_next_month),
    }

    if filter_type == 'overdue':
        return Stage.objects.filter(
            project__is_archived=False,
            status__in=["Not started", "In Progress"],
            planned_date__lt=today
        ).order_by('planned_date')
    elif filter_type in date_ranges:
        start, end = date_ranges[filter_type]
        return Stage.objects.filter(
            project__is_archived=False,
            status__in=["Not started", "In Progress"],
            planned_date__range=(start, end)
        ).order_by('planned_date')
    elif filter_type == 'all':
        return Stage.objects.filter(project__is_archived=False).exclude(status__in=["Completed", "Not Applicable"]).order_by('planned_date')
    else:
        return Stage.objects.filter(
            project__is_archived=False,
            status__in=["Not started", "In Progress"],
            planned_date__gte=today
        ).order_by('planned_date')
    
@login_required
def export_milestones_excel(request):
    filter_type = request.GET.get('filter', 'all').capitalize()
    stages = get_filtered_stages(filter_type)

    timestamp = datetime.now().strftime('%d-%m-%Y %H:%M')
    filename = f'Upcoming Milestones {filter_type} {timestamp}.csv'

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['Project Code', 'Customer', 'Milestone', 'Status', 'Planned Date'])

    for stage in stages:
        writer.writerow([
            stage.project.code,
            stage.project.customer_name,
            stage.name,
            stage.status,
            stage.planned_date
        ])
    return response

@login_required
def export_milestones_pdf(request):
    raw_filter = request.GET.get('filter', 'all')
    filter_type = raw_filter.lower()
    stages = get_filtered_stages(filter_type)

    # Format filename as "Upcoming Milestones [Filter] [dd-mm-yyyy HH-MM].pdf"
    timestamp = datetime.now().strftime('%d-%m-%Y %H:%M')
    filename = f'Upcoming Milestones {filter_type.capitalize()} {timestamp}.pdf'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph(f"Upcoming Milestones - Filter: {filter_type.capitalize()}", styles['Heading2']))

    data = [['Project Code', 'Customer', 'Milestone', 'Status', 'Planned Date']]
    for stage in stages:
        data.append([
            stage.project.code,
            stage.project.customer_name,
            stage.name,
            stage.status,
            stage.planned_date.strftime('%Y-%m-%d') if stage.planned_date else 'N/A'
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _build_filtered_report_projects(request):
    """Applies the Reports page's standard filters (segments, team leads, value range,
    financial year, stages, planned & actual date windows) to the Project queryset.
    Shared by the PDF and Excel exports so their filtering stays in lockstep with the
    on-page report."""
    query_params = request.GET
    today = timezone.now().date()
    plausible_planned_date_floor = today - timedelta(days=MAX_PLAUSIBLE_DAY_DELTA)

    projects_qs = Project.objects.filter(is_archived=False).select_related('segment_con', 'team_lead').prefetch_related('stages').all()

    selected_segment_ids = query_params.getlist('segments')
    selected_team_lead_ids = query_params.getlist('team_leads')
    selected_stage_keys = query_params.getlist('stages')
    min_value = query_params.get('min_value')
    max_value = query_params.get('max_value')
    selected_fy = query_params.get('financial_year')

    actual_start_date = parse_date(query_params.get('actual_start_date')) if query_params.get('actual_start_date') else None
    actual_end_date = parse_date(query_params.get('actual_end_date')) if query_params.get('actual_end_date') else None

    period = query_params.get('period')
    if not period:
        period = 'custom' if (query_params.get('planned_start_date') and query_params.get('planned_end_date')) else 'this_month'

    def _month_range(d):
        month_start = d.replace(day=1)
        month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
        return month_start, month_end

    def _quarter_range(d):
        q_start_month = ((d.month - 1) // 3) * 3 + 1
        q_start = date(d.year, q_start_month, 1)
        q_end = (q_start + relativedelta(months=3)) - timedelta(days=1)
        return q_start, q_end

    if period == 'this_quarter':
        start_date, end_date = _quarter_range(today)
    elif period == 'all':
        start_date, end_date = date(2000, 1, 1), today
    elif period == 'custom':
        start_date = parse_date(query_params.get('planned_start_date')) if query_params.get('planned_start_date') else None
        end_date = parse_date(query_params.get('planned_end_date')) if query_params.get('planned_end_date') else None
        if not (start_date and end_date):
            period = 'this_month'
            start_date, end_date = _month_range(today)
    else:
        period = 'this_month'
        start_date, end_date = _month_range(today)

    if selected_fy:
        year = int(selected_fy)
        fy_start = date(year, 4, 1)
        fy_end = date(year + 1, 3, 31)
        start_date, end_date = fy_start, fy_end

        completed_early_ids = Project.objects.filter(
            is_archived=False, stages__name='Handover', stages__status='Completed', stages__actual_date__lt=fy_start
        ).values_list('id', flat=True)
        projects_qs = projects_qs.filter(
            Q(so_punch_date__lte=fy_end) | Q(so_punch_date__isnull=True)
        ).exclude(id__in=completed_early_ids)
        if selected_stage_keys:
            projects_qs = projects_qs.filter(stages__name__in=selected_stage_keys)
    else:
        planned_window_stages = Stage.objects.filter(
            planned_date__isnull=False, planned_date__gte=plausible_planned_date_floor,
            planned_date__lte=end_date,
        ).exclude(Q(status='Not Applicable') | (Q(status='Completed') & Q(actual_date__lt=start_date)))
        if selected_stage_keys:
            planned_window_stages = planned_window_stages.filter(name__in=selected_stage_keys)
        projects_qs = projects_qs.filter(id__in=planned_window_stages.values_list('project_id', flat=True))

    if actual_start_date and actual_end_date:
        actual_window_stages = Stage.objects.filter(
            status='Completed', actual_date__range=[actual_start_date, actual_end_date],
        )
        if selected_stage_keys:
            actual_window_stages = actual_window_stages.filter(name__in=selected_stage_keys)
        projects_qs = projects_qs.filter(id__in=actual_window_stages.values_list('project_id', flat=True))

    if selected_segment_ids:
        projects_qs = projects_qs.filter(segment_con__id__in=selected_segment_ids)
    if selected_team_lead_ids:
        projects_qs = projects_qs.filter(team_lead__id__in=selected_team_lead_ids)
    if min_value:
        try: projects_qs = projects_qs.filter(value__gte=float(min_value))
        except (ValueError, TypeError): pass
    if max_value:
        try: projects_qs = projects_qs.filter(value__lte=float(max_value))
        except (ValueError, TypeError): pass

    if query_params.get('hide_completed') == '1':
        completed_project_ids = Project.objects.filter(
            is_archived=False, stages__name='Handover', stages__status='Completed'
        ).values_list('id', flat=True)
        projects_qs = projects_qs.exclude(id__in=completed_project_ids)

    stage_names_to_report = (
        [(k, v) for k, v in Stage.STAGE_NAMES if k in selected_stage_keys] if selected_stage_keys
        else Stage.STAGE_NAMES
    )

    return {
        'projects_qs': projects_qs.distinct(),
        'today': today,
        'start_date': start_date,
        'end_date': end_date,
        'stage_names_to_report': stage_names_to_report,
    }


@login_required
def export_report_pdf(request):
    filtered = _build_filtered_report_projects(request)
    distinct_projects = filtered['projects_qs']

    # --- Start Building the PDF ---
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f"Project_Report_{timestamp}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Filtered Project Report", styles['Title']))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}", styles['Normal']))
    
    # Table Data
    table_data = [['Code', 'Customer', 'Segment', 'Value (INR)', 'Completion %', 'Status']]
    for p in distinct_projects:
        table_data.append([
            p.code,
            p.customer_name,
            p.segment_con.name if p.segment_con else 'N/A',
            f"{p.value:,.2f}",
            f"{p.get_completion_percentage()}%",
            p.get_overall_status()
        ])


    project_table = Table(table_data)
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))


    elements.append(project_table)
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


@login_required
def export_report_excel(request):
    """Excel version of the Reports page: full filter fidelity (segments, team leads,
    financial year, stages, planned & actual date windows, value range, hide completed),
    plus a Stage Summary sheet mirroring the on-page planned/actual/backlog table."""
    filtered = _build_filtered_report_projects(request)
    distinct_projects = filtered['projects_qs']
    today = filtered['today']
    summary_start = filtered['start_date']
    summary_end = filtered['end_date']
    stage_names_to_report = filtered['stage_names_to_report']
    chart_project_ids = list(distinct_projects.values_list('id', flat=True))
    plausible_planned_date_floor = today - timedelta(days=MAX_PLAUSIBLE_DAY_DELTA)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f"Project_Report_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = Workbook()

    # --- Sheet 1: Projects ---
    projects_sheet = workbook.active
    projects_sheet.title = "Projects"
    projects_sheet.append(['Code', 'Customer', 'Segment', 'Team Lead', 'Value (Cr)', 'SO Punch Date', 'Completion %', 'Status', 'OTIF %'])
    for cell in projects_sheet[1]:
        cell.font = Font(bold=True)

    for p in distinct_projects:
        projects_sheet.append([
            p.code,
            p.customer_name,
            p.segment_con.name if p.segment_con else '',
            p.team_lead.name if p.team_lead else '',
            float(p.value) if p.value is not None else None,
            p.so_punch_date.strftime('%Y-%m-%d') if p.so_punch_date else '',
            p.get_completion_percentage(),
            p.get_overall_status(),
            p.get_otif_percentage(),
        ])

    # --- Sheet 2: Stage Summary (mirrors the on-page Planned vs Actual table) ---
    summary_sheet = workbook.create_sheet("Stage Summary")
    summary_sheet.append([
        'Stage', 'Type', 'Planned', 'Actual Completed', 'Pending (Backlog)',
        'Currently Delayed', 'Avg Delay (days)', 'Avg Cycle Time (days)', 'OTIF %'
    ])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

    automation_stage_keys = dict(Stage.AUTOMATION_STAGES)
    for stage_key, stage_display in stage_names_to_report:
        planned_backlog_qs = Stage.objects.filter(
            project_id__in=chart_project_ids, name=stage_key,
            planned_date__isnull=False, planned_date__gte=plausible_planned_date_floor,
            planned_date__lte=summary_end,
        ).exclude(Q(status='Not Applicable') | (Q(status='Completed') & Q(actual_date__lt=summary_start)))

        actual_period_qs = Stage.objects.filter(
            project_id__in=chart_project_ids, name=stage_key, status='Completed',
            actual_date__range=[summary_start, summary_end]
        )

        delayed_qs = planned_backlog_qs.filter(status__in=['Not started', 'In Progress'], planned_date__lt=today)

        planned_count = planned_backlog_qs.count()
        pending_count = planned_backlog_qs.exclude(status='Completed').count()
        delayed_count = delayed_qs.count()
        actual_count = actual_period_qs.count()
        on_time_count = actual_period_qs.filter(actual_date__lte=F('planned_date')).count()
        otif_pct = round((on_time_count / actual_count) * 100, 1) if actual_count else None

        overdue_planned_dates = list(delayed_qs.values_list('planned_date', flat=True))
        overdue_deltas = [(today - d).days for d in overdue_planned_dates if 0 <= (today - d).days <= MAX_PLAUSIBLE_DAY_DELTA]
        avg_delay_days = round(sum(overdue_deltas) / len(overdue_deltas), 1) if overdue_deltas else None

        cycle_pairs = list(actual_period_qs.exclude(planned_start_date__isnull=True).values_list('planned_start_date', 'actual_date'))
        cycle_deltas = [(a - p).days for p, a in cycle_pairs if 0 <= (a - p).days <= MAX_PLAUSIBLE_DAY_DELTA]
        avg_cycle_time = round(sum(cycle_deltas) / len(cycle_deltas), 1) if cycle_deltas else None

        summary_sheet.append([
            stage_display,
            'Automation' if stage_key in automation_stage_keys else 'Emulation',
            planned_count, actual_count, pending_count, delayed_count,
            avg_delay_days if avg_delay_days is not None else '',
            avg_cycle_time if avg_cycle_time is not None else '',
            otif_pct if otif_pct is not None else '',
        ])

    for sheet in (projects_sheet, summary_sheet):
        for col_cells in sheet.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 40)

    workbook.save(response)
    return response


@login_required
def add_remark(request, stage_id):
    stage = get_object_or_404(Stage, id=stage_id)
    if request.method == 'POST':
        text = request.POST.get('remark')
        if text:
            StageRemark.objects.create(stage=stage, text=text, added_by=request.user)
            messages.success(request, "Remark added.")
    return redirect('tracker_project_detail', project_id=stage.project.id)

@login_required
def edit_project_comment(request, comment_id):
    comment = get_object_or_404(ProjectComment, id=comment_id)
    if request.user == comment.added_by or request.user.is_staff:
        if request.method == 'POST':
            new_text = request.POST.get('note_text')
            if new_text:
                comment.text = new_text
                comment.save()
                # messages.success(request, "Note updated successfully.")
    else:
        messages.error(request, "You do not have permission to edit this note.")
    
    return HttpResponseRedirect(f"{reverse('tracker_project_detail', args=[comment.project.id])}#project-notes")

@login_required
def delete_project_comment(request, comment_id):
    comment = get_object_or_404(ProjectComment, id=comment_id)
    project_id = comment.project.id
    if request.user == comment.added_by or request.user.is_staff:
        comment.delete()
        # messages.success(request, "Note deleted successfully.")
    else:
        messages.error(request, "You do not have permission to delete this note.")
    
    return HttpResponseRedirect(f"{reverse('tracker_project_detail', args=[project_id])}#project-notes")

@login_required
def get_remarks(request, stage_id):
    stage = get_object_or_404(Stage, id=stage_id)
    return render(request, 'tracker/view_remarks_modal.html', {'stage': stage})

@login_required
def add_project_update(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    if request.method == 'POST':
        text = request.POST.get('update_text')
        push_pull_type = request.POST.get('push_pull_type')
        who_contact_ids = request.POST.getlist('who_contact')
        raised_by_id = request.POST.get('raised_by')
        eta = parse_date(request.POST.get('eta_date')) if request.POST.get('eta_date') else None

        if text and push_pull_type:
            update = ProjectUpdate.objects.create(
                project=project,
                author=request.user,
                text=text,
                push_pull_type=push_pull_type,
                eta=eta,
                raised_by_id=raised_by_id if raised_by_id else None,
                content_type='Project',
            )

            for contact_id in who_contact_ids:
                if contact_id:
                    try:
                        contact = ContactPerson.objects.get(pk=contact_id)
                        update.who_contact.add(contact)
                    except ContactPerson.DoesNotExist:
                        pass

            # messages.success(request, "Push-Pull content added successfully.")
        else:
            messages.error(request, "Update text and type are required.")
    
    return HttpResponseRedirect(f"{reverse('tracker_project_detail', args=[project.id])}?bottom_tab=push_pull#project-notes")

@login_required
def add_general_update(request):
    if request.method == 'POST':
        text = request.POST.get('update_text')
        push_pull_type = request.POST.get('push_pull_type')
        who_contact_ids = request.POST.getlist('who_contact')
        raised_by_id = request.POST.get('raised_by')
        project_id = request.POST.get('project_id')

        eta = parse_date(request.POST.get('eta_date')) if request.POST.get('eta_date') else None

        if text and push_pull_type:
            content_type = 'General'
            project = None
            
            if project_id and project_id != 'general':
                try:
                    project = Project.objects.get(pk=project_id)
                    content_type = 'Project'
                except Project.DoesNotExist:
                    pass

            update = ProjectUpdate.objects.create(
                author=request.user,
                text=text,
                push_pull_type=push_pull_type,
                eta=eta,
                raised_by_id=raised_by_id if raised_by_id else None,
                content_type=content_type,
                project=project,
            )
            for contact_id in who_contact_ids:
                if contact_id:
                    try:
                        contact = ContactPerson.objects.get(pk=contact_id)
                        update.who_contact.add(contact)
                    except ContactPerson.DoesNotExist:
                        pass

            messages.success(request, "Content added successfully.")
        else:
            messages.error(request, "Update text and type are required.")

    return redirect('all_push_pull_content')


@login_required
def edit_project_update(request, update_id):
    update = get_object_or_404(ProjectUpdate, id=update_id)
    if request.user == update.author or request.user.groups.filter(name='Trackers').exists():
        if request.method == 'POST':
            new_status = request.POST.get('update_status', update.status)
            
            if new_status == 'Closed' and update.status != 'Closed':
                update.closed_at = timezone.now()
            elif new_status != 'Closed' and update.status == 'Closed':
                update.closed_at = None
            
            # Update text and push_pull_type only if they're in the POST data
            if 'update_text' in request.POST:
                update.text = request.POST.get('update_text', update.text)
            if 'push_pull_type' in request.POST:
                update.push_pull_type = request.POST.get('push_pull_type', update.push_pull_type)
            
            update.status = new_status
            
            # Only update ETA if it's in the POST data
            if 'eta_date' in request.POST:
                eta_date = request.POST.get('eta_date')
                update.eta = parse_date(eta_date) if eta_date else None
            
            update.save()
            
            # Only update who_contact if it's in the POST data
            if 'who_contact' in request.POST:
                who_contact_ids = request.POST.getlist('who_contact')
                update.who_contact.clear()
                for contact_id in who_contact_ids:
                    if contact_id:
                        try:
                            contact = ContactPerson.objects.get(pk=contact_id)
                            update.who_contact.add(contact)
                        except ContactPerson.DoesNotExist:
                            pass

            # messages.success(request, "Push-Pull content saved successfully.")
        
        referer = request.META.get('HTTP_REFERER')
        if referer and 'all-push-pull-content' in referer:
            return redirect('all_push_pull_content')
            
        if update.project:
            return HttpResponseRedirect(f"{reverse('tracker_project_detail', args=[update.project.id])}?bottom_tab=push_pull#project-notes")
        else:
            return redirect('all_push_pull_content')
    else:
        messages.error(request, "You do not have permission to edit this update.")
        if update.project:
            return redirect('tracker_project_detail', project_id=update.project.id)
        else:
            return redirect('all_push_pull_content')


@login_required
def delete_project_update(request, update_id):
    update = get_object_or_404(ProjectUpdate, id=update_id)
    if request.user == update.author or request.user.is_staff:
        project_id = update.project.id if update.project else None
        update.delete()
        messages.success(request, "Push-Pull content deleted.")
        referer = request.META.get('HTTP_REFERER')
        if referer and 'all-push-pull-content' in referer:
            return redirect('all_push_pull_content')
        
        if project_id:
            return HttpResponseRedirect(f"{reverse('tracker_project_detail', args=[project_id])}?bottom_tab=push_pull#project-notes")
        else:
            return redirect('all_push_pull_content')
    else:
        messages.error(request, "You do not have permission to delete this update.")
        if update.project:
            return redirect('tracker_project_detail', project_id=update.project.id)
        else:
            return redirect('all_push_pull_content')

@login_required
def toggle_update_status(request, update_id):
    update = get_object_or_404(ProjectUpdate, id=update_id)
    if update.status == 'Open':
        update.status = 'Closed'
    else:
        update.status = 'Open'
    update.save()
    messages.success(request, f"'{update.category}' status changed to {update.status}.")
    return redirect('tracker_project_detail', project_id=update.project.id)


@login_required
def save_mitigation_plan(request, update_id):
    update = get_object_or_404(ProjectUpdate, id=update_id)
    if request.method == 'POST' and update.category == 'Risk':
        update.mitigation_plan = request.POST.get('mitigation_plan')
        update.save()
        messages.success(request, "Mitigation plan saved.")
    return redirect('tracker_project_detail', project_id=update.project.id)

@login_required
def all_project_updates(request, project_id):
    project = get_object_or_404(Project, pk=project_id)


    updates = project.updates.select_related('author', 'raised_by').prefetch_related('who_contact', 'remarks__added_by').all()
    contact_persons = ContactPerson.objects.all()


    context = {
        'project': project,
        'updates': updates,
        'contact_persons': contact_persons,

    }
    return render(request, 'tracker/all_project_updates.html', context)


@login_required
def all_push_pull_content(request, filter=None):
    # Handle explicit clear filters request
    if request.GET.get('clear_filters') == '1':
        for key in ['pp_category_filter', 'pp_status_filter', 'pp_type_filter']:
            if key in request.session:
                del request.session[key]
        return redirect('all_push_pull_content')

    # If filter is passed as part of the URL path (e.g., from a redirect kwarg)
    if filter:
        request.session['pp_category_filter'] = filter

    # Session-based filter persistence from query parameters
    if 'filter' in request.GET:
        request.session['pp_category_filter'] = request.GET['filter']
    if 'status_filter' in request.GET:
        request.session['pp_status_filter'] = request.GET['status_filter']
    if 'push_pull_filter' in request.GET:
        request.session['pp_type_filter'] = request.GET['push_pull_filter']

    # Get the current filters from the session (default to 'all')
    current_filter = request.session.get('pp_category_filter', 'all')
    status_filter = request.session.get('pp_status_filter', 'all')
    push_pull_filter = request.session.get('pp_type_filter', 'all')

    # Auto-archive logic: Move 'Closed' items older than 30 days to 'Archived'
    archive_threshold = timezone.now() - timedelta(days=30)
    ProjectUpdate.objects.filter(status='Closed', closed_at__lt=archive_threshold).update(status='Archived')



    updates_qs = ProjectUpdate.objects.select_related('author', 'project', 'raised_by').prefetch_related('who_contact', 'remarks__added_by').exclude(project__is_archived=True).order_by('-created_at')

    if current_filter == 'project':
        updates_qs = updates_qs.filter(content_type='Project')
    elif current_filter == 'general':
        updates_qs = updates_qs.filter(content_type='General')

    # ✅ NEW: Apply push/pull filtering
    if push_pull_filter == 'push':
        updates_qs = updates_qs.filter(push_pull_type='Push')
    elif push_pull_filter == 'pull':
        updates_qs = updates_qs.filter(push_pull_type='Pull')

    # Apply status filtering
    if status_filter == 'open':
        updates_qs = updates_qs.exclude(status__in=['Closed', 'Archived'])
    elif status_filter == 'closed':
        updates_qs = updates_qs.filter(status='Closed')
    elif status_filter == 'archived':
        updates_qs = updates_qs.filter(status='Archived')
    else: # 'all'
        updates_qs = updates_qs.exclude(status='Archived')

    updates = list(updates_qs.all())
    for update in updates:
        for remark in update.remarks.all():
            if remark.added_by:
                remark.added_by.username = remark.added_by.get_full_name() or remark.added_by.username
    contact_persons = ContactPerson.objects.all()
    projects = Project.objects.all()

    context = {
        'updates': updates,
        'contact_persons': contact_persons,
        'projects': projects,
        'filter': current_filter,

        'status_filter': status_filter,
        'push_pull_filter': push_pull_filter, # ✅ NEW: Pass push/pull filter to template

    }
    return render(request, 'tracker/all_push_pull_content.html', context)

@login_required
def add_contact_person_ajax(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        name = request.POST.get('name')
        if not name:
            return JsonResponse({'status': 'error', 'message': 'Name is required'}, status=400)
        
        # ✅ NEW: Validate that the name has at least a first and last name
        parts = name.strip().split(' ', 1)
        if len(parts) < 2 or not parts[0] or not parts[1]:
            return JsonResponse({'status': 'error', 'message': 'First and last name are required.'}, status=400)

        
        contact, created = ContactPerson.objects.get_or_create(name=name)
        
        return JsonResponse({'status': 'success', 'id': contact.id, 'name': contact.name, 'created': created})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=405)


@login_required
def export_push_pull_excel(request):
    updates_qs = ProjectUpdate.objects.select_related('project', 'author', 'raised_by').prefetch_related('who_contact', 'remarks__added_by').exclude(project__is_archived=True).order_by('-created_at')
    filter = request.GET.get('filter', request.session.get('pp_category_filter', 'all'))

    if filter == 'project':
        updates_qs = updates_qs.filter(content_type='Project')
    elif filter == 'general':
        updates_qs = updates_qs.filter(content_type='General')

    # Apply push/pull filtering
    push_pull_filter = request.GET.get('push_pull_filter', request.session.get('pp_type_filter', 'all'))
    if push_pull_filter == 'push':
        updates_qs = updates_qs.filter(push_pull_type='Push')
    elif push_pull_filter == 'pull':
        updates_qs = updates_qs.filter(push_pull_type='Pull')

    # Apply status filtering
    status_filter = request.GET.get('status_filter', request.session.get('pp_status_filter', 'all'))
    if status_filter == 'open':
        updates_qs = updates_qs.exclude(status__in=['Closed', 'Archived'])
    elif status_filter == 'closed':
        updates_qs = updates_qs.filter(status='Closed')
    elif status_filter == 'archived':
        updates_qs = updates_qs.filter(status='Archived')
    else: # 'all' or None
        updates_qs = updates_qs.exclude(status='Archived')

    updates = updates_qs.all()
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"all_push_pull_contents_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "All Push-Pull Contents"
    
    # Headers
    headers = [
        'Project Code', 'Type', 'What', 'Who', 'Raised By', 'ETA', 'Status', 'Created At', 'Closed At', 'Remarks'
    ]
    sheet.append(headers)
    
    # Populate with data
    for update in updates:
        # Join multiple contacts with a comma
        who_contacts_str = ", ".join([p.name for p in update.who_contact.all()])
        
        remarks_list = []
        for r in update.remarks.all():
            user_str = (r.added_by.get_full_name() or r.added_by.username) if r.added_by else "Unknown"
            remarks_list.append(f"{user_str} ({r.created_at.strftime('%Y-%m-%d %H:%M')}): {r.text}")
        remarks_text = " | ".join(remarks_list)

        # --- FIX START ---
        # A date object does not have a tzinfo attribute, so it doesn't need to be replaced.
        # A datetime object does, and must be made naive to be compatible with openpyxl.
        created_at_naive = update.created_at.replace(tzinfo=None) if update.created_at else None
        closed_at_naive = update.closed_at.replace(tzinfo=None) if update.closed_at else None
        
        # The ETA field is a DateField, so no timezone removal is needed.
        eta_naive = update.eta
        # --- FIX END ---

        row = [
            update.project.code if update.project else 'N/A',
            update.get_push_pull_type_display(),
            update.text,
            who_contacts_str,
            update.raised_by.name if update.raised_by else '-',
            eta_naive,
            update.status,
            created_at_naive,
            closed_at_naive,
            remarks_text
        ]
        sheet.append(row)

    workbook.save(response)
    
    return response

@login_required
def export_push_pull_pdf(request):
    updates_qs = ProjectUpdate.objects.select_related('project', 'author', 'raised_by').prefetch_related('who_contact', 'remarks__added_by').exclude(project__is_archived=True).order_by('-created_at')
    filter = request.GET.get('filter', request.session.get('pp_category_filter', 'all'))

    if filter == 'project':
        updates_qs = updates_qs.filter(content_type='Project')
    elif filter == 'general':
        updates_qs = updates_qs.filter(content_type='General')

    # Apply push/pull filtering
    push_pull_filter = request.GET.get('push_pull_filter', request.session.get('pp_type_filter', 'all'))
    if push_pull_filter == 'push':
        updates_qs = updates_qs.filter(push_pull_type='Push')
    elif push_pull_filter == 'pull':
        updates_qs = updates_qs.filter(push_pull_type='Pull')

    # Apply status filtering
    status_filter = request.GET.get('status_filter', request.session.get('pp_status_filter', 'all'))
    if status_filter == 'open':
        updates_qs = updates_qs.exclude(status__in=['Closed', 'Archived'])
    elif status_filter == 'closed':
        updates_qs = updates_qs.filter(status='Closed')
    elif status_filter == 'archived':
        updates_qs = updates_qs.filter(status='Archived')
    else: # 'all' or None
        updates_qs = updates_qs.exclude(status='Archived')

    updates = updates_qs.all()

    buffer = BytesIO()
    # Use landscape A4 for higher width capacity in a list-style report
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles for a cleaner report
    body_style = ParagraphStyle(
        'body_style',
        parent=styles['Normal'],
        spaceBefore=6,
        spaceAfter=6,
        leading=14,
    )
    section_header_style = ParagraphStyle(
        'section_header_style',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        spaceBefore=12,
        spaceAfter=2,
    )

    elements.append(Paragraph("All Push-Pull Contents", styles['Title']))
    elements.append(Paragraph(f"Report Generated on: {timezone.now().strftime('%d-%b-%Y %I:%M %p')}", styles['Normal']))
    elements.append(Spacer(1, 1*cm))

    for update in updates:
        who_contacts_str = ", ".join([p.name for p in update.who_contact.all()])

        # Create a small table for the metadata of each update
        details_data = [
            [Paragraph('<b>Project:</b>', styles['Normal']), Paragraph(escape(update.project.code if update.project else 'General'), styles['Normal'])],
            [Paragraph('<b>Type:</b>', styles['Normal']), Paragraph(update.get_push_pull_type_display(), styles['Normal'])],
            [Paragraph('<b>Status:</b>', styles['Normal']), Paragraph(update.status, styles['Normal'])],
            [Paragraph('<b>ETA:</b>', styles['Normal']), Paragraph(update.eta.strftime('%Y-%m-%d') if update.eta else '-', styles['Normal'])],
            [Paragraph('<b>Raised By:</b>', styles['Normal']), Paragraph(escape(update.raised_by.name if update.raised_by else '-'), styles['Normal'])],
            [Paragraph('<b>Who:</b>', styles['Normal']), Paragraph(escape(who_contacts_str) if who_contacts_str else '-', body_style)],
        ]
        details_table = Table(details_data, colWidths=[3*cm, None])
        details_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('LEFTPADDING', (0, 0), (-1, -1), 4), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(details_table)

        # Add the "What" description as a separate, splittable Paragraph
        elements.append(Paragraph('What (Description):', section_header_style))
        what_text = escape(update.text).replace('\n', '<br/>') if update.text else "-"
        elements.append(Paragraph(what_text, body_style))

        # Add Remarks as a series of splittable Paragraphs
        elements.append(Paragraph('Remarks:', section_header_style))
        remarks = list(update.remarks.order_by('created_at'))
        if not remarks:
            elements.append(Paragraph('-', body_style))
        else:
            for r in remarks:
                user_str = (r.added_by.get_full_name() or r.added_by.username) if r.added_by else "Unknown"
                safe_remark_text = escape(r.text).replace('\n', '<br/>')
                safe_user_str = escape(user_str)
                remark_str = f"• <b>{safe_user_str}</b> ({r.created_at.strftime('%b %d, %H:%M')}): {safe_remark_text}"
                elements.append(Paragraph(remark_str, body_style))

        # Add a separator before the next update
        elements.append(Spacer(1, 1*cm))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"all_push_pull_contents_{timezone.now().strftime('%Y-%m-%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response

@login_required
def send_push_pull_email(request):
    # 1. Retrieve filters (similar to all_push_pull_content)
    # Check GET first, then session, default to 'all'
    current_filter = request.GET.get('filter', request.session.get('pp_category_filter', 'all'))
    status_filter = request.GET.get('status_filter', request.session.get('pp_status_filter', 'all'))
    push_pull_filter = request.GET.get('push_pull_filter', request.session.get('pp_type_filter', 'all'))

    # 2. Filter the updates
    updates_qs = ProjectUpdate.objects.select_related('author', 'project', 'raised_by').prefetch_related('who_contact', 'remarks__added_by').exclude(project__is_archived=True).order_by('-created_at')

    if current_filter == 'project':
        updates_qs = updates_qs.filter(content_type='Project')
    elif current_filter == 'general':
        updates_qs = updates_qs.filter(content_type='General')

    if push_pull_filter == 'push':
        updates_qs = updates_qs.filter(push_pull_type='Push')
    elif push_pull_filter == 'pull':
        updates_qs = updates_qs.filter(push_pull_type='Pull')

    if status_filter == 'open':
        updates_qs = updates_qs.exclude(status__in=['Closed', 'Archived'])
    elif status_filter == 'closed':
        updates_qs = updates_qs.filter(status='Closed')
    elif status_filter == 'archived':
        updates_qs = updates_qs.filter(status='Archived')
    else:
        updates_qs = updates_qs.exclude(status='Archived')

    updates = list(updates_qs.all())

    # 3. Collect recipients from the 'who_contact' field
    recipients = set()
    for update in updates:
        for contact in update.who_contact.all():
            if contact.email:
                recipients.add(contact.email)

    if not recipients:
        messages.error(request, "No email addresses found for the contacts in the current list.")
        return redirect('all_push_pull_content')

    # 4. Generate Public Link (using the hardcoded token from public view)
    token = "a1b2c3d4-e5f6-7890-1234-567890abcdef"
    public_url = request.build_absolute_uri(reverse('public_push_pull_content', args=[token]))
    
    # Append filters to the public URL so the recipient sees the same view
    params = request.GET.copy()
    if 'filter' not in params and current_filter != 'all':
        params['filter'] = current_filter
    if params:
        public_url += f"?{params.urlencode()}"

    # 5. Render the Email Body using the new dedicated template
    context = {
        'updates': updates,
        'public_url': public_url,
        'date_str': timezone.now().strftime('%d-%b-%Y'),
    }

    try:
        # Render the dedicated email template
        html_content = render_to_string('tracker/email_push_pull_content.html', context, request=request)

        send_mail(
            subject=f"Push/Pull Content Update - {timezone.now().strftime('%d-%b-%Y')}",
            message="", # Plain text fallback left empty as we rely on HTML
            from_email=None, # Use DEFAULT_FROM_EMAIL
            recipient_list=list(recipients),
            html_message=html_content,
            fail_silently=False
        )
        messages.success(request, f"Email sent successfully to {len(recipients)} recipients.")
    except Exception as e:
        messages.error(request, f"Error sending email: {e}")

    # Redirect back preserving filters
    redirect_url = reverse('all_push_pull_content')
    if request.GET:
        redirect_url += f"?{request.GET.urlencode()}"
    return HttpResponseRedirect(redirect_url)


@login_required
def help_page(request):
    """
    Renders the help and documentation page.
    """
    return render(request, 'tracker/help_page.html')

@login_required
def update_stage_ajax(request, stage_id):
    if request.method == 'POST':
        stage = get_object_or_404(Stage, id=stage_id)
        
        try:
            data = json.loads(request.body)
            field_name = data.get('field_name')
            new_value = data.get('new_value')

            old_value = getattr(stage, field_name)

            # Map field names to user-friendly names for the history log
            field_map = {
                'planned_start_date': 'Planned Start Date',
                'planned_date': 'Planned Finish Date',
                'status': 'Status',
                'actual_date': 'Actual Finish Date',
                'completion_percentage': '% Completion',
            }
            history_field_name = field_map.get(field_name, field_name.replace('_', ' ').title())

            # Update the field on the stage object
            fields_to_update = [field_name]
            if 'date' in field_name:
                setattr(stage, field_name, parse_date(new_value) if new_value else None)
            else:
                setattr(stage, field_name, new_value)

            updated_actual_date = None

            if field_name == 'status':
                if new_value == 'Completed' and not stage.actual_date:
                    stage.actual_date = timezone.now().date()
                    updated_actual_date = stage.actual_date.strftime('%Y-%m-%d')
                    fields_to_update.append('actual_date')
                elif new_value != 'Completed':
                    stage.actual_date = None
                    updated_actual_date = '' # Signal frontend to clear it
                    fields_to_update.append('actual_date')
            
            # Prevent clearing actual date if status is Completed
            if field_name == 'actual_date' and not new_value and stage.status == 'Completed':
                 return JsonResponse({'status': 'error', 'message': 'Cannot remove Actual Date while status is Completed.'}, status=400)
            
            # Use update_fields to prevent race conditions (e.g. Status vs % Completion updates)
            stage.save(update_fields=fields_to_update)

            # Create a history record of the change
            StageHistory.objects.create(
                stage=stage,
                changed_by=request.user,
                field_name=history_field_name,
                old_value=str(old_value),
                new_value=str(new_value)
            )
            
            response_data = {'status': 'success', 'message': 'Stage updated successfully.'}
            if updated_actual_date is not None:
                response_data['updated_actual_date'] = updated_actual_date
            
            return JsonResponse(response_data)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

@login_required
def update_project_update_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            update_id = data.get('id')
            field = data.get('field')
            value = data.get('value')

            update = get_object_or_404(ProjectUpdate, id=update_id)

            # Permission check: Author, Trackers group, or Staff
            if not (request.user == update.author or request.user.groups.filter(name='Trackers').exists() or request.user.is_staff):
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

            if field == 'text':
                update.text = value
            elif field == 'push_pull_type':
                update.push_pull_type = value
            elif field == 'status':
                if value == 'Closed' and update.status != 'Closed':
                    update.closed_at = timezone.now()
                elif value != 'Closed':
                    update.closed_at = None
                update.status = value
            elif field == 'eta':
                update.eta = parse_date(value) if value else None
            elif field == 'who_contact':
                # Value is expected to be a list of IDs from Select2
                update.who_contact.clear()
                if isinstance(value, list):
                    for contact_id in value:
                        if contact_id:
                            try:
                                contact = ContactPerson.objects.get(id=contact_id)
                                update.who_contact.add(contact)
                            except ContactPerson.DoesNotExist:
                                pass
            elif field == 'raised_by':
                if value:
                    try:
                        contact = ContactPerson.objects.get(id=value)
                        update.raised_by = contact
                    except ContactPerson.DoesNotExist:
                        pass
                else:
                    update.raised_by = None

            update.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@login_required
def add_update_remark(request, update_id):
    update = get_object_or_404(ProjectUpdate, id=update_id)
    if request.method == 'POST':
        text = request.POST.get('remark_text')
        redirect_to = request.POST.get('redirect_to') # ✅ Get the new hidden field
        if text:
            remark = UpdateRemark.objects.create(
                update=update,
                text=text,
                added_by=request.user
            )

            messages.success(request, "Remark added successfully.")
            
    # ✅ Corrected redirect logic to handle both project and non-project updates
    if redirect_to == 'project_detail' and update.project:
        # Redirect back to the specific project detail page on the push-pull tab
        return HttpResponseRedirect(f"{reverse('tracker_project_detail', args=[update.project.id])}?bottom_tab=push_pull#project-notes")
    else:
        # Default to the general content page or the filtered page
        return redirect('all_push_pull_content')

@login_required
def edit_update_remark(request, remark_id):
    remark = get_object_or_404(UpdateRemark, pk=remark_id)
    redirect_to = None

    if request.method == 'POST':
        # Grab the redirect_to value from the form first
        redirect_to = request.POST.get('redirect_to')
        new_text = request.POST.get('remark_text')

        # Security check: only the author or a staff member can edit
        if request.user == remark.added_by or request.user.is_staff:
            if new_text:
                remark.text = new_text
                remark.save()
                messages.success(request, "Remark updated successfully.")
        else:
            messages.error(request, "You do not have permission to edit this remark.")
    
    # ✅ Corrected redirect logic
    if redirect_to == 'project_detail' and remark.update.project:
        # Redirect back to the project page on the push-pull tab
        return HttpResponseRedirect(f"{reverse('tracker_project_detail', args=[remark.update.project.id])}?bottom_tab=push_pull#project-notes")
    else:
        # Default to the all push-pull content page
        return redirect('all_push_pull_content')


@login_required
def delete_update_remark(request, remark_id):
    remark = get_object_or_404(UpdateRemark, pk=remark_id)
    redirect_to = request.POST.get('redirect_to')
    
    # Security check: only the author or a staff member can delete
    if request.user == remark.added_by or request.user.is_staff:
        if request.method == 'POST':
            remark.delete()
            messages.success(request, "Remark deleted successfully.")
    else:
        messages.error(request, "You do not have permission to delete this remark.")


    # ✅ Corrected redirect logic
    if redirect_to == 'project_detail' and remark.update.project:
        # Redirect back to the project page on the push-pull tab if the update has a project
        return HttpResponseRedirect(f"{reverse('tracker_project_detail', args=[remark.update.project.id])}?bottom_tab=push_pull#project-notes")
    else:
        # Redirect back to the general content page
        return redirect('all_push_pull_content')
    

from uuid import UUID

def public_push_pull_content(request, access_token):
    try:
        # The access_token is ALREADY a UUID object because of the URL converter.
        valid_token = UUID("a1b2c3d4-e5f6-7890-1234-567890abcdef") 

        if access_token != valid_token:
            return HttpResponse("Unauthorized", status=401)
            
    except ValueError:
        return HttpResponse("Invalid Token", status=400)
    
    # ✅ NEW: Read filter parameters from the URL
    current_filter = request.GET.get('filter', 'all')
    status_filter = request.GET.get('status_filter', 'all')
    push_pull_filter = request.GET.get('push_pull_filter', 'all')

    updates_qs = ProjectUpdate.objects.select_related('author', 'project', 'raised_by').prefetch_related('who_contact', 'remarks', 'remarks__added_by').exclude(project__is_archived=True).order_by('-created_at')

    if current_filter == 'project':
        updates_qs = updates_qs.filter(content_type='Project')
    elif current_filter == 'general':
        updates_qs = updates_qs.filter(content_type='General')

    if push_pull_filter == 'push':
        updates_qs = updates_qs.filter(push_pull_type='Push')
    elif push_pull_filter == 'pull':
        updates_qs = updates_qs.filter(push_pull_type='Pull')
        
    if status_filter == 'open':
        updates_qs = updates_qs.exclude(status__in=['Closed', 'Archived'])
    elif status_filter == 'closed':
        updates_qs = updates_qs.filter(status='Closed')
    elif status_filter == 'archived':
        updates_qs = updates_qs.filter(status='Archived')
    else:
        updates_qs = updates_qs.exclude(status='Archived')

    updates = list(updates_qs.all())
    for update in updates:
        for remark in update.remarks.all():
            if remark.added_by:
                remark.added_by.username = remark.added_by.get_full_name() or remark.added_by.username
    contact_persons = ContactPerson.objects.all()

    # The redirect logic needs to be updated to redirect back to the public URL
    if request.method == 'POST':
        update_id = request.POST.get('update_id')
        update = get_object_or_404(ProjectUpdate, id=update_id)

        # Build query string to retain filters for the public view
        from urllib.parse import urlencode
        q_string = request.GET.urlencode()
        redirect_url = reverse('public_push_pull_content', args=[str(access_token)])
        if q_string:
            redirect_url += f"?{q_string}"

        if 'update_status' in request.POST:
            update.status = request.POST['update_status']
            update.save()
            messages.success(request, f"Update status for item {update.id} changed to {update.status}.")
            return redirect(redirect_url)
        
        if 'remark_text' in request.POST:
            text = request.POST.get('remark_text')
            if text:
                try:
                    public_user = get_user_model().objects.get(username='public_user')
                except get_user_model().DoesNotExist:
                    public_user = get_user_model().objects.create_user('public_user', 'public@example.com', 'some_random_password', is_active=False)

                UpdateRemark.objects.create(
                    update=update,
                    text=text,
                    added_by=public_user
                )
                messages.success(request, "Remark added successfully.")
            return redirect(redirect_url)
    
    context = {
        'updates': updates,
        'contact_persons': contact_persons,
        'is_public_view': True,
        'access_token': str(access_token),
        'filter': current_filter,
        'status_filter': status_filter,
        'push_pull_filter': push_pull_filter,
    }
    return render(request, 'tracker/all_push_pull_content.html', context)
    
    return render(request, 'tracker/all_push_pull_content.html', context)