"""
Excel workbook export for the Reports & Analytics page.

Every section shown on the page becomes a sheet here, plus native Excel charts
(status distribution, category distribution, inventory-by-type) - this is the
"everything, with graphs" export, as distinct from the plain items-only CSV
still offered on the Items list page.
"""
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone

HEADER_FILL = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
SECTION_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)


def _write_table(ws, start_row, headers, rows):
    """Writes a header row + data rows starting at start_row. Returns the row after the table."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    for r, row_data in enumerate(rows, start=start_row + 1):
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=r, column=col, value=value)
    return start_row + len(rows) + 1


def _autosize(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _who(user):
    if not user:
        return ''
    return user.get_full_name() or user.username


def _build_summary_sheet(ws, data):
    ws.title = 'Summary'
    ws['A1'] = 'Inventory Reports & Analytics'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"

    row = 4
    ws.cell(row=row, column=1, value='Key Metrics').font = SECTION_FONT
    row += 1
    metrics = [
        ('Total Inventory Value', float(data['total_value'])),
        ('Overdue Assignments', data['summary_counts']['overdue_assignments']),
        ('Overdue Dispatches', data['summary_counts']['overdue_dispatches']),
        ('Low Stock Materials', data['summary_counts']['low_stock_items']),
        ('Under Maintenance', data['summary_counts']['maintenance_items']),
        ('Pending Reservations', data['reservation_summary']['pending']),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if label == 'Total Inventory Value':
            cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Inventory Summary by Type').font = SECTION_FONT
    row += 1
    table_start = row
    headers = ['Item Type', 'Total', 'Available', 'Assigned', 'Dispatched', 'Maintenance', 'Consumed', 'Retired']
    rows = [
        [s['item_type'], s['total'], s['available'], s['assigned'], s['dispatched'],
         s['maintenance'], s['consumed'], s['retired']]
        for s in data['inventory_summary']
    ]
    next_row = _write_table(ws, table_start, headers, rows)

    if rows:
        chart = BarChart()
        chart.title = 'Inventory by Status per Type'
        chart.y_axis.title = 'Count'
        chart_data = Reference(ws, min_col=3, max_col=8, min_row=table_start, max_row=table_start + len(rows))
        cats = Reference(ws, min_col=1, min_row=table_start + 1, max_row=table_start + len(rows))
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 18
        ws.add_chart(chart, f'A{next_row + 1}')

    row = next_row + 20
    ws.cell(row=row, column=1, value='Return Condition Breakdown (all-time returns)').font = SECTION_FONT
    row += 1
    breakdown = data['return_condition_breakdown']
    condition_rows = [
        ['Good', breakdown.get('GOOD', 0)],
        ['Damaged', breakdown.get('DAMAGED', 0)],
        ['Needs Repair', breakdown.get('NEEDS_REPAIR', 0)],
        ['Lost', breakdown.get('LOST', 0)],
    ]
    _write_table(ws, row, ['Condition', 'Count'], condition_rows)

    _autosize(ws, [26, 14, 14, 14, 14, 14, 14, 14])


def _build_items_sheet(ws, items):
    ws.title = 'Items'
    headers = ['Name', 'Type', 'Serial Number', 'Make', 'Model', 'Status', 'Quantity',
               'Min Quantity', 'Location', 'Category', 'Purchase Date', 'Purchase Cost',
               'Current Value', 'Remarks']
    rows = []
    for item in items:
        current_value = float(item.purchase_cost) * item.quantity if item.purchase_cost is not None else None
        rows.append([
            item.name, item.get_item_type_display(), item.serial_number, item.make, item.model,
            item.get_status_display(), item.quantity, item.min_quantity, item.location, item.category,
            item.purchase_date.isoformat() if item.purchase_date else None,
            float(item.purchase_cost) if item.purchase_cost is not None else None,
            current_value, item.remarks,
        ])
    _write_table(ws, 1, headers, rows)
    _autosize(ws, [24, 10, 18, 14, 14, 14, 10, 12, 18, 14, 14, 14, 14, 34])


def _build_distribution_sheet(ws, data):
    ws.title = 'Distribution'
    ws.cell(row=1, column=1, value='Status Distribution (all item types combined)').font = SECTION_FONT

    totals = {
        'Available': sum(s['available'] for s in data['inventory_summary']),
        'Assigned': sum(s['assigned'] for s in data['inventory_summary']),
        'Dispatched': sum(s['dispatched'] for s in data['inventory_summary']),
        'Maintenance': sum(s['maintenance'] for s in data['inventory_summary']),
        'Consumed': sum(s['consumed'] for s in data['inventory_summary']),
        'Retired': sum(s['retired'] for s in data['inventory_summary']),
    }
    status_start = 3
    status_rows = [[label, count] for label, count in totals.items()]
    _write_table(ws, status_start, ['Status', 'Count'], status_rows)

    if any(count for _, count in status_rows):
        pie = PieChart()
        pie.title = 'Status Distribution'
        chart_data = Reference(ws, min_col=2, min_row=status_start, max_row=status_start + len(status_rows))
        cats = Reference(ws, min_col=1, min_row=status_start + 1, max_row=status_start + len(status_rows))
        pie.add_data(chart_data, titles_from_data=True)
        pie.set_categories(cats)
        pie.height = 9
        pie.width = 12
        ws.add_chart(pie, f'D{status_start}')

    _autosize(ws, [22, 12, 14])


def _build_reservations_sheet(ws, data):
    ws.title = 'Reservations'
    ws.cell(row=1, column=1, value='Reservation Summary').font = SECTION_FONT
    rs = data['reservation_summary']
    summary_rows = [
        ['Pending', rs['pending']],
        ['Fulfilled', rs['fulfilled']],
        ['Cancelled', rs['cancelled']],
        ['Expired', rs['expired']],
        ['Total', rs['total']],
    ]
    next_row = _write_table(ws, 3, ['Status', 'Count'], summary_rows)

    row = next_row + 2
    ws.cell(row=row, column=1, value='Upcoming Reservations (next 30 days)').font = SECTION_FONT
    row += 1
    upcoming_rows = [
        [r.item.name, r.item.serial_number, _who(r.reserved_for), r.start_date.isoformat(), r.end_date.isoformat()]
        for r in data['upcoming_reservations']
    ]
    _write_table(ws, row, ['Tool', 'Serial Number', 'Reserved For', 'Start Date', 'End Date'], upcoming_rows)
    _autosize(ws, [24, 18, 26, 14, 14])


def _build_assigned_sheet(ws, data):
    ws.title = 'Assigned Items'
    headers = ['Item', 'Serial Number', 'Assigned To', 'Since', 'Expected Return', 'Overdue', 'Days Overdue']
    rows = [
        [a.item.name, a.item.serial_number, _who(a.assigned_to),
         a.assignment_date.isoformat(),
         a.expected_return_date.isoformat() if a.expected_return_date else '',
         'Yes' if a.is_overdue else 'No', a.days_overdue]
        for a in data['assigned_items']
    ]
    _write_table(ws, 1, headers, rows)
    _autosize(ws, [24, 18, 26, 14, 16, 10, 14])


def _build_dispatches_sheet(ws, data):
    ws.title = 'Active Dispatches'
    headers = ['Item', 'Serial Number', 'Project', 'Site Location', 'Responsible Person',
               'Since', 'Overdue', 'Days Overdue']
    rows = [
        [d.item.name, d.item.serial_number, d.project, d.site_location or '', _who(d.responsible_person),
         d.dispatch_date.isoformat(), 'Yes' if d.is_overdue else 'No', d.days_overdue]
        for d in data['active_dispatches']
    ]
    _write_table(ws, 1, headers, rows)
    _autosize(ws, [24, 18, 20, 18, 26, 14, 10, 14])


def _build_low_stock_sheet(ws, data):
    ws.title = 'Low Stock'
    headers = ['Material', 'Serial Number', 'Current Stock', 'Reorder Threshold']
    rows = [[i.name, i.serial_number, i.quantity, i.min_quantity] for i in data['low_stock_items']]
    _write_table(ws, 1, headers, rows)
    _autosize(ws, [24, 18, 16, 18])


def _build_user_assignments_sheet(ws, data):
    ws.title = 'User Assignments'
    headers = ['User', 'Email', 'Tools Assigned', 'Last Assignment Date', 'Has Overdue']
    rows = []
    for entry in data['user_assignments']:
        last = entry['last_assignment']
        rows.append([
            _who(entry['user']), entry['user'].email, entry['tool_count'],
            last.assignment_date.isoformat() if last else '',
            'Yes' if entry['has_overdue'] else 'No',
        ])
    _write_table(ws, 1, headers, rows)
    _autosize(ws, [26, 32, 14, 18, 12])


def build_reports_workbook(data, items):
    """
    `data` is the dict returned by inventory.utils.build_reports_context_data().
    `items` is an iterable of Item for the full item-detail sheet.
    Returns an openpyxl Workbook, ready to .save(response).
    """
    wb = Workbook()
    _build_summary_sheet(wb.active, data)
    _build_items_sheet(wb.create_sheet(), items)
    _build_distribution_sheet(wb.create_sheet(), data)
    _build_reservations_sheet(wb.create_sheet(), data)
    _build_assigned_sheet(wb.create_sheet(), data)
    _build_dispatches_sheet(wb.create_sheet(), data)
    _build_low_stock_sheet(wb.create_sheet(), data)
    _build_user_assignments_sheet(wb.create_sheet(), data)
    return wb
